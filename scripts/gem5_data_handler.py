import csv
import re
import subprocess
from collections import defaultdict

import openpyxl
from pandas import read_csv
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Border, Side, Font
# # 根据列的数字返回字母
# print(get_column_letter(2))  # B
# # 根据字母返回列的数字
# print(column_index_from_string('D'))  # 4


import time


# pip3 install openpyxl

#
# 抓取gem5 Ruby Cache l2size 不同变化的数据，汇总成表格
#

def runcmd(command):
    ret = subprocess.run(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, encoding="utf-8",
                         timeout=None)
    if ret.returncode == 0:
        # print("success:", ret)
        return ret.stdout.split('\n')
    else:
        # print("error:", ret)
        return ret.stderr.split('\n')


def grep_data(grep_keys_list, gem5_stats_path_list):
    results_dict = defaultdict(defaultdict)
    for f in gem5_stats_path_list:
        print(f)
        with open(f, 'r', encoding='utf-8') as fr:
            for line in fr.readlines():
                for index, k in enumerate(grep_keys_list):
                    group = re.search(
                        r'(?P<itemName>' + k[
                            0] + r')\s+' + r'(?P<itemValue>(\-|\+)?\d+(\.\d+)?)\s+(?P<itemComments>#.*)',
                        line
                    )
                    if group:
                        results_dict[f][k[1]] = group["itemValue"]
    return results_dict


def set_gem5_stats_path_list_with_ruby_l2_change():
    gem5_stats_path_list = []
    gem5_stats_path_list.extend(
        runcmd(["find /home/lizongping/Desktop/gem5spec/verification24/*/*r/ -name 'gem5_stats.log'"]))
    gem5_stats_path_list.extend(
        runcmd(["find /home/lizongping/Desktop/gem5spec/verification21/*/*r/ -name 'gem5_stats.log'"]))
    gem5_stats_path_list.extend(
        runcmd(["find /home/lizongping/Desktop/gem5spec/verification20/*/*r/ -name 'gem5_stats.log'"]))
    gem5_stats_path_list.extend(
        runcmd(["find /home/lizongping/Desktop/gem5spec/verification25/*/*r/ -name 'gem5_stats.log'"]))

    # gem5_stats_path_list.extend(
    #     runcmd(["find ../data/home/lizongping/Desktop/gem5spec/verification21/*/*r/ -name 'gem5_stats.log'"]))
    gem5_stats_path_list = list(sorted(set(gem5_stats_path_list)))
    if "" in gem5_stats_path_list:
        gem5_stats_path_list.remove("")
    if '\n' in gem5_stats_path_list:
        gem5_stats_path_list.remove('\n')
    # [print(f) for f in gem5_stats_path_list]
    return gem5_stats_path_list


def set_grep_keys_with_ruby_l2_change():
    # "data,alias"
    return [
        ["system.cpu.ipc", "IPC"],
        ["system.cpu.committedInsts", "Insts"],
        ["system.cpu.numCycles", "Cycles"],
        ["system.ruby.l1_cntrl0.L1Icache.m_demand_hits", "I Cache Hits"],
        ["system.ruby.l1_cntrl0.L1Icache.m_demand_misses", "I Cache Misses"],
        ["system.ruby.l1_cntrl0.L1Icache.m_demand_accesses", "I Cache Accesses"],
        ["system.ruby.l1_cntrl0.L1Dcache.m_demand_hits", "D Cache Hits"],
        ["system.ruby.l1_cntrl0.L1Dcache.m_demand_misses", "D Cache Misses"],
        ["system.ruby.l1_cntrl0.L1Dcache.m_demand_accesses", "D Cache Accesses"],
        ["system.ruby.network.ext_links1.ext_node.L2cache.m_demand_hits", "L2 Cache Hits"],
        ["system.ruby.network.ext_links1.ext_node.L2cache.m_demand_misses", "L2 Cache Misses"],
        ["system.ruby.network.ext_links1.ext_node.L2cache.m_demand_accesses", "L2 Cache Accesses"],
        ["system.ruby.l3_cntrl0.L3cache.m_demand_hits", "L3 Cache Hits"],
        ["system.ruby.l3_cntrl0.L3cache.m_demand_misses", "L3 Cache Misses"],
        ["system.ruby.l3_cntrl0.L3cache.m_demand_accesses", "L3 Cache Accesses"],
        ["system.cpu.commit.loads", "system.cpu.commit.loads"]
    ]


# 自动设置列宽
def auto_set_column_width(worksheet):
    dims = {}
    for row in worksheet.rows:
        for cell in row:
            if cell.value:
                # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                # 大写字母的长度近似多0.45
                cell_len = 0.7 * len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(
                    str(cell.value)) + 0.45 * len(re.findall('([A-Z])', str(cell.value)))
                # print(float(str(cell.font.size)))
                # print(cell.column)
                if cell.font.size:
                    if float(cell.font.size) > 11:
                        cell_len += float(cell.font.size) - 11
                if cell.font.bold:
                    cell_len += 1
                dims[cell.column] = max((dims.setdefault(cell.column, 0), cell_len))
    for col, value in dims.items():
        # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+2是用来调整最终效果的
        worksheet.column_dimensions[get_column_letter(col)].width = value + 1
        # print(col,value)


# 改变字体的样式,大小,粗体,斜体;wss表示[sheet..]
def pyCellFontStyle(wss, onlyRow=False, rowNumStart: int = 1, onlyCol=False,
                    colNumStart: int = 1, rowAndCol=False, fontName="Times New Roman", fontSize=11, isbold=False,
                    italic=False):
    if onlyRow:
        for ws in wss:
            nrows = ws.max_row  # 获得行数
            ncols = ws.max_column
            for j in range(ncols):
                # name代表样式，size代表大小，bold代表粗体，italic代表斜体
                ws.cell(rowNumStart, j + 1).font = Font(name=fontName, size=fontSize, bold=isbold, italic=italic)
    if onlyCol:
        for ws in wss:
            nrows = ws.max_row  # 获得行数
            ncols = ws.max_column
            for i in range(nrows):
                # name代表样式，size代表大小，bold代表粗体，italic代表斜体
                ws.cell(i + 1, colNumStart).font = Font(name=fontName, size=fontSize, bold=isbold, italic=italic)
    if rowAndCol:
        for ws in wss:
            nrows = ws.max_row  # 获得行数
            ncols = ws.max_column
            for i in range(nrows):
                for j in range(ncols):
                    # name代表样式，size代表大小，bold代表粗体，italic代表斜体
                    ws.cell(rowNumStart + i, colNumStart + j).font = Font(name=fontName, size=fontSize,
                                                                          bold=isbold, italic=italic)


def set_line_chart_mark(chart,color=None, symbol="circle", line_dashStyle="solid",line_noFill=False):
    for index, serie in enumerate(chart.series):
        serie.marker.symbol = symbol  # 标记样式
        # {‘triangle’, ‘dash’, ‘dot’, ‘star’, ‘circle’, ‘picture’, ‘square’, ‘x’, ‘plus’, ‘auto’, ‘diamond’}
        if color:
            serie.marker.graphicalProperties.solidFill = color  # 标记 Marker filling
            serie.marker.graphicalProperties.line.solidFill = color  # Marker outline
            serie.graphicalProperties.line.solidFill = color  # 线条颜色
        serie.graphicalProperties.line.noFill = line_noFill  # 有没有线条,False默认有线条
        serie.graphicalProperties.line.dashStyle = line_dashStyle  # 线条样式


def add_desc(workbook: openpyxl.Workbook):
    data = {
        "specTest.beginRunTime": "2022-08-01",
        "specTest.lastGetTime": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
        "specTest.testName": "verification20/21/22/23/24/25",
        "specTest.gem5spec_path": "/home/lizongping/Desktop/gem5spec",
        "specTest.testPath": "/home/lizongping/Desktop/gem5spec/verification20/21/22/23/24/25",
        "specTest.SPEC_HOME": "/home/lizongping/cpu2017",
        "specTest.LABEL": "ppc",
        "specTest.GEM5_REPO_PATH": "/home/lizongping/Desktop/gitlab/gem5_L2test",
        "specTest.GEM5": "$(GEM5_REPO_PATH)/build/POWER_MI/gem5.opt",
        "specTest.mem_size": "--mem-size=16384MB",
        "specTest.cpu_type": "--cpu-type=O3CPU",
        "specTest.caches": "--ruby --caches",
        "specTest.l1d_size": "--l1d_size=64kB",
        "specTest.l1d_assoc": "--l1d_assoc=8",
        "specTest.l1i_size": "--l1i_size=32kB",
        "specTest.l1i_assoc": "--l1i_assoc=8",
        "specTest.l2cache": "--l2cache",
        "specTest.l2_size": "--l2_size=128kB / --l2_size=256kB / --l2_size=512kB / --l2_size=1MB / --l2_size=2MB / --l2_size=4MB",
        "specTest.l2_assoc": "--l2_assoc=8",
        "specTest.l3_size": "--l3_size=8MB",
        "specTest.l3_assoc": "--l3_assoc=8",
        "specTest.cacheline_size": "--cacheline_size=128",
    }
    sheet = workbook.create_sheet("Configs", index=None)
    # sheet.title = "Config"
    for k, v in data.items():
        sheet.append([k, v])
    pyCellFontStyle([sheet], rowAndCol=True, rowNumStart=1, colNumStart=1, fontName="Times New Roman", fontSize=11)
    pyCellFontStyle([sheet], onlyRow=True, rowNumStart=18, fontName="Times New Roman", fontSize=11, isbold=True)
    auto_set_column_width(sheet)


# 生成csv
def gen_csv_with_ruby_l2_change():
    localtime = time.strftime("%Y%m%d", time.localtime())
    # print(str(localtime))
    results_dict = grep_data(set_grep_keys_with_ruby_l2_change(), set_gem5_stats_path_list_with_ruby_l2_change())
    bm_metric_dict: dict[str, dict[str, list[str]]] = {}
    l2_switch_case = {"128k": 0, "256k": 1, "512k": 2, "1M": 3, "2M": 4, "4M": 5}
    gen_file = localtime + "_P8_Ruby_L2_Cache"
    with open("../data/" + gen_file + ".csv", "w+", encoding="utf-8") as fw:
        p8_ruby_l2_cache_writer = csv.writer(fw)
        header = ["Benchmark", "Metrics", "L2=128k", "L2=256k", "L2=512k", "L2=1M", "L2=2M", "L2=4M"]
        p8_ruby_l2_cache_writer.writerow(header)
        for f, v1 in results_dict.items():
            group = re.search(
                r'/ppc_MI_Three_Level_L2_(?P<L2Size>\d+[k|M])/(?P<BM>[\w.]+)/', f
            )
            l2size = group["L2Size"]
            bm = group["BM"]
            for metric, v2 in v1.items():
                bm_metric_dict.setdefault(bm, {}).setdefault(metric, [None] * 6)[l2_switch_case[l2size]] = v2
                # print(metric,v2)
        for bm, v1 in bm_metric_dict.items():
            # for metric, v2 in v1.items():
            for metric in set_grep_keys_with_ruby_l2_change():  # 保持原顺序
                v2 = v1[metric[1]]
                v2.insert(0, metric[1])
                v2.insert(0, bm)
                p8_ruby_l2_cache_writer.writerow(v2)
    # csv->excel
    with open("../data/" + gen_file + ".csv", "r", encoding="utf-8") as fr:
        read_csv(fr).to_excel("../data/" + gen_file + ".xlsx")

    # excel plot
    workbook = load_workbook(filename="../data/" + gen_file + ".xlsx")
    sheet = workbook["Sheet1"]
    sheet.delete_cols(1)  # 删除原默认行号

    # print(sheet.max_row)
    sheet.title = "Metrics"
    pyCellFontStyle([sheet], onlyRow=True, rowNumStart=1, fontName="Times New Roman", fontSize=12, isbold=True)
    pyCellFontStyle([sheet], rowAndCol=True, rowNumStart=2, colNumStart=1, fontName="Times New Roman", fontSize=11)
    # 计算 cache miss rate
    start_row = 2
    # IPC总图
    chart3_IPC = LineChart()
    chart3_IPC.title = "IPC Trends"
    chart3_IPC.width = 18
    chart3_IPC.height = 18
    chart4_Insts = LineChart()
    chart4_Insts.title = "Insts Trends"
    chart4_Insts.width = 18
    chart4_Insts.height = 18
    chart5_Cycles = LineChart()
    chart5_Cycles.title = "Cycles Trends"
    chart5_Cycles.width = 18
    chart5_Cycles.height = 18
    while start_row < sheet.max_row:
        sheet.insert_rows(start_row + 12)
        sheet.insert_rows(start_row + 12)
        sheet.insert_rows(start_row + 9)
        sheet.insert_rows(start_row + 9)
        sheet.insert_rows(start_row + 6)
        sheet.insert_rows(start_row + 6)
        sheet.insert_rows(start_row + 3)
        sheet.insert_rows(start_row + 3)
        # insert metrics
        sheet['B' + str(start_row + 3)] = "I Cache Miss Rate"
        sheet['B' + str(start_row + 4)] = "I Cache Miss / 1000 Insts"
        sheet['B' + str(start_row + 8)] = "D Cache Miss Rate"
        sheet['B' + str(start_row + 9)] = "D Cache Miss / 1000 Insts"
        sheet['B' + str(start_row + 13)] = "L2 Cache Miss Rate"
        sheet['B' + str(start_row + 14)] = "L2 Cache Miss / 1000 Insts"
        sheet['B' + str(start_row + 18)] = "L3 Cache Miss Rate"
        sheet['B' + str(start_row + 19)] = "L3 Cache Miss / 1000 Insts"

        # 计算每个benchmark的miss rate和千条miss次数
        for c in range(column_index_from_string('C'), column_index_from_string('H') + 1):
            column = get_column_letter(c)
            # I Cache Miss Rate  &  I Cache Miss / 1000 Insts
            sheet[column + str(start_row + 3)] = "=" + column + str(start_row + 6) + "/" + column + str(start_row + 7)
            sheet[column + str(start_row + 4)] = "=1000*" + column + str(start_row + 6) + "/" + column + str(
                start_row + 1)

            sheet[column + str(start_row + 8)] = "=" + column + str(start_row + 11) + "/" + column + str(start_row + 12)
            sheet[column + str(start_row + 9)] = "=1000*" + column + str(start_row + 11) + "/" + column + str(
                start_row + 1)

            sheet[column + str(start_row + 13)] = "=" + column + str(start_row + 16) + "/" + column + str(
                start_row + 17)
            sheet[column + str(start_row + 14)] = "=1000*" + column + str(start_row + 16) + "/" + column + str(
                start_row + 1)

            sheet[column + str(start_row + 18)] = "=" + column + str(start_row + 21) + "/" + column + str(
                start_row + 22)
            sheet[column + str(start_row + 19)] = "=1000*" + column + str(start_row + 21) + "/" + column + str(
                start_row + 1)
        # 每千条指令的CacheMiss次数
        chart1 = LineChart()
        # CacheMiss Rate
        chart2 = LineChart()

        chart1.add_data(Reference(sheet, min_col=column_index_from_string('B'), max_col=column_index_from_string('H'),
                                  min_row=start_row + 4, max_row=start_row + 4), from_rows=True, titles_from_data=True)
        chart1.add_data(Reference(sheet, min_col=column_index_from_string('B'), max_col=column_index_from_string('H'),
                                  min_row=start_row + 9, max_row=start_row + 9), from_rows=True, titles_from_data=True)
        chart1.add_data(Reference(sheet, min_col=column_index_from_string('B'), max_col=column_index_from_string('H'),
                                  min_row=start_row + 14, max_row=start_row + 14), from_rows=True,
                        titles_from_data=True)
        chart1.add_data(Reference(sheet, min_col=column_index_from_string('B'), max_col=column_index_from_string('H'),
                                  min_row=start_row + 19, max_row=start_row + 19), from_rows=True,
                        titles_from_data=True)
        # set x-axis values
        chart1.set_categories(Reference(sheet, min_row=1, max_row=1, min_col=column_index_from_string('C'),
                                        max_col=column_index_from_string('H')))

        chart2.add_data(Reference(sheet, min_col=column_index_from_string('B'), max_col=column_index_from_string('H'),
                                  min_row=start_row + 3, max_row=start_row + 3), from_rows=True, titles_from_data=True)
        chart2.add_data(Reference(sheet, min_col=column_index_from_string('B'), max_col=column_index_from_string('H'),
                                  min_row=start_row + 8, max_row=start_row + 8), from_rows=True, titles_from_data=True)
        chart2.add_data(Reference(sheet, min_col=column_index_from_string('B'), max_col=column_index_from_string('H'),
                                  min_row=start_row + 13, max_row=start_row + 13), from_rows=True,
                        titles_from_data=True)
        chart2.add_data(Reference(sheet, min_col=column_index_from_string('B'), max_col=column_index_from_string('H'),
                                  min_row=start_row + 18, max_row=start_row + 18), from_rows=True,
                        titles_from_data=True)
        # set x-axis values
        chart2.set_categories(Reference(sheet, min_row=1, max_row=1, min_col=column_index_from_string('C'),
                                        max_col=column_index_from_string('H')))
        chart3_values = Reference(sheet, min_col=column_index_from_string('C'), max_col=column_index_from_string('H'),
                                  min_row=start_row, max_row=start_row)
        chart3_series = Series(chart3_values, title=sheet['A' + str(start_row)].value)
        chart3_IPC.append(chart3_series)

        chart4_values = Reference(sheet, min_col=column_index_from_string('C'), max_col=column_index_from_string('H'),
                                  min_row=start_row + 1, max_row=start_row + 1)
        chart4_series = Series(chart4_values, title=sheet['A' + str(start_row)].value)
        chart4_Insts.append(chart4_series)

        chart5_values = Reference(sheet, min_col=column_index_from_string('C'), max_col=column_index_from_string('H'),
                                  min_row=start_row + 2, max_row=start_row + 2)
        chart5_series = Series(chart5_values, title=sheet['A' + str(start_row)].value)
        chart5_Cycles.append(chart5_series)
        # 设置第一列居中
        area = 'A' + str(start_row) + ":" + 'A' + str(start_row + 23)
        sheet.merge_cells(range_string=area)
        for i in sheet[area]:
            for j in i:
                j.alignment = Alignment(horizontal='center', vertical='center')
        chart1.title = str(sheet['A' + str(start_row)].value) + " Cache Miss Per 1000 Insts"
        chart1.width = 20
        chart2.title = str(sheet['A' + str(start_row)].value) + " Cache Miss Rate"
        chart2.width = 18

        color = ["6495ED", "EE7942", "696969", "EEC900"]
        # 设置折线图标记
        for index, serie in enumerate(chart1.series):
            serie.marker.symbol = "circle"  # 标记样式
            serie.marker.graphicalProperties.solidFill = color[index]  # 标记 Marker filling
            serie.marker.graphicalProperties.line.solidFill = color[index]  # Marker outline
            serie.graphicalProperties.line.noFill = False
            # serie.graphicalProperties.line.dashStyle = "sysDot" # 线条样式
            # print(serie.graphicalProperties.line.dashStyle)
            serie.graphicalProperties.line.solidFill = color[index]  # 线条颜色

        for index, serie in enumerate(chart2.series):
            serie.marker.symbol = "circle"  # 标记样式
            serie.marker.graphicalProperties.solidFill = color[index]  # 标记 Marker filling
            serie.marker.graphicalProperties.line.solidFill = color[index]  # Marker outline
            serie.graphicalProperties.line.noFill = False
            # serie.graphicalProperties.line.dashStyle = "sysDot" # 线条样式
            serie.graphicalProperties.line.solidFill = color[index]  # 线条颜色
        # 添加横纵坐标说明
        chart1.x_axis.title = "Cache Size"
        chart1.y_axis.title = "Cache Miss Number"
        chart2.x_axis.title = "Cache Size"
        chart2.y_axis.title = "Cache Miss Rate"
        sheet.add_chart(chart1, "J" + str(start_row))  # 在工作表上添加图表，并指定图表左上角锚定的单元格。
        sheet.add_chart(chart2, "W" + str(start_row))  # 在工作表上添加图表，并指定图表左上角锚定的单元格。
        # 设置border分隔每个benchmark数据
        green = "00008000"
        double = Side(border_style="double", color=green)
        area = 'A' + str(start_row + 23) + ":" + "AG" + str(start_row + 23)
        for i in sheet[area]:
            for j in i:
                j.border = Border(bottom=double)
        # 设置关键数据加粗
        # for c in range(column_index_from_string('B'), column_index_from_string('H') + 1):
        #     column = get_column_letter(c)
        #     area = [column + str(start_row + 4), column + str(start_row + 9), column + str(start_row + 14),
        #             column + str(start_row + 19)]
        #     for a in area:
        #         sheet[a].font = Font(bold=True)
        area = [start_row + 4, start_row + 9, start_row + 14, start_row + 19]
        for a in area:
            # sheet[a].font = Font(bold=True)
            pyCellFontStyle([sheet], onlyRow=True, rowNumStart=a, isbold=True)
        start_row += 24
        # print(sheet['A1'].font.size)
        workbook.save("../data/" + gen_file + ".xlsx")

    # set x-axis values
    chart3_IPC.set_categories(Reference(sheet, min_row=1, max_row=1, min_col=column_index_from_string('C'),
                                        max_col=column_index_from_string('H')))
    chart4_Insts.set_categories(Reference(sheet, min_row=1, max_row=1, min_col=column_index_from_string('C'),
                                          max_col=column_index_from_string('H')))
    chart5_Cycles.set_categories(Reference(sheet, min_row=1, max_row=1, min_col=column_index_from_string('C'),
                                           max_col=column_index_from_string('H')))
    chart3_IPC.x_axis.title = "Cache Size"
    chart3_IPC.y_axis.title = "IPC Number"
    set_line_chart_mark(chart3_IPC,symbol="auto")
    sheet.add_chart(chart3_IPC, "A" + str(start_row))  # 在工作表上添加图表，并指定图表左上角锚定的单元格。

    chart4_Insts.x_axis.title = "Cache Size"
    chart4_Insts.y_axis.title = "Insts Number"
    set_line_chart_mark(chart4_Insts,symbol="auto")
    sheet.add_chart(chart4_Insts, "Q" + str(start_row))  # 在工作表上添加图表，并指定图表左上角锚定的单元格。

    chart5_Cycles.x_axis.title = "Cache Size"
    chart5_Cycles.y_axis.title = "Cycles Number"
    set_line_chart_mark(chart5_Cycles,symbol="auto")
    sheet.add_chart(chart5_Cycles, "G" + str(start_row))  # 在工作表上添加图表，并指定图表左上角锚定的单元格。

    auto_set_column_width(sheet)
    add_desc(workbook)
    workbook.save("../data/" + gen_file + ".xlsx")


gen_csv_with_ruby_l2_change()
