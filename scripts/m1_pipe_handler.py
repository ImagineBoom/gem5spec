import csv
import datetime
import os
import subprocess
import re
import threading
from collections import namedtuple
from concurrent.futures import ProcessPoolExecutor
from concurrent.futures import ThreadPoolExecutor
from concurrent.futures import Future
from os import getpid as pid
from os.path import basename
from time import sleep

from openpyxl.reader.excel import load_workbook


#
# 在scripts目录下执行
#
# ../data/pipeline_result/ 保存最终结果
# ../data/pipeline_graph/ 保存展开折叠后的所有流水线图文件
# ../data/meta/scripts_csv_power-isa-implementation.csv 是依赖的源文件，用做检索的key

def runcmd(command):
    ret = subprocess.run(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, encoding="utf-8",
                         timeout=None)
    if ret.returncode == 0:
        # print("success:", ret)
        return ret.stdout.split('\n')
    else:
        # print("error:", ret)
        return ret.stderr.split('\n')


class ExeCycle:
    def __init__(self, cycle_num=-1, cycle_frequency=0.0, cycle_count=0):
        self.cycle_num = cycle_num  # 执行周期数,-1代表当前片段未执行完该指令，无法统计指令执行周期
        self.cycle_frequency = cycle_frequency  # 对应执行周期数出现的频率
        self.cycle_count = cycle_count  # 对应执行周期数出现的次数


class IssueCycle:
    def __init__(self, cycle_num=-1, cycle_frequency=0.0, cycle_count=0):
        self.cycle_num = cycle_num  # 周期数,-1代表当前片段未执行完该指令，无法统计指令周期
        self.cycle_frequency = cycle_frequency  # 对应周期数出现的频率
        self.cycle_count = cycle_count  # 对应周期数出现的次数


# 具体的每条指令
class instruction:
    def __init__(self, IopId, MnemonicValue, InstAddrLow32, DataAddrLow32, Pipe, filename="") -> None:
        self.IopId: int = IopId
        self.MnemonicValue: str = MnemonicValue
        self.InstAddrLow32: str = InstAddrLow32
        self.DataAddrLow32: str = DataAddrLow32
        self.exeCycle = ExeCycle()
        self.issueCycle = IssueCycle()
        self.Pipe = Pipe
        self.filename = filename
        self.benchmark = ""
        self.location = -1 # 当前Instruction是第几条

    def set_benchmark(self,benchmark_name):
        self.benchmark = benchmark_name

    def set_location(self,instruction_location):
        self.location = instruction_location

# 每种(助记符相同的)指令
class Instruction:
    def __init__(self, count=0, frequency=0.0):
        self.list: list[instruction] = []
        self.count = count
        self.frequency = frequency
        self.exeCycle: dict[int,ExeCycle] = {}
        self.issueCycle: dict[int,IssueCycle] = {}


class Trace:
    def __init__(self, fname, existed_files):
        # super(Trace, self).__init__(name=fname)
        self.fname: str = fname
        self.instruction_dict: dict[str, Instruction] = {}
        self.inst_count = 0  # 所有遍历的指令数
        self.lines: list[str] = []  # 流水线图的行
        self.cur_line = ""  # 当前处理的流水线图的行
        self.pattern = re.compile(
            r'^\|(?P<pipe>.*)[\+\-\|]+\s+(?P<iop_id>\d+)\s*\|\s*(?P<mnemonic_key>[\w\?\-]+)\s*(?P<mnemonic_value>[\w\,\.\(\)\?\+\-]*)\s*\|\s*(?P<inst_addr_low32>\w+)\s*\|\s*(?P<data_addr_low32>\w*)\s*\|$')
        self.existed_files = existed_files

    def m1pipe_read(self, fname: str):
        with open(fname, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        return lines

    def fwrite(self, fname: str):
        with open(fname, 'w', encoding='utf-8') as f:
            f.write('test')

    def m1pipe_grep(self, fname: str):
        self.lines = self.m1pipe_read(fname)
        start = False
        for line in self.lines:
            if "| Scrollpipe Data |" in line:
                start = True
                # print(self.name + "-started")
                # print(line)
            if not start:
                # print("not start")
                continue
            # print("start")
            pipeline = self.pattern.search(line)
            if pipeline:
                # print(pipeline.group(0))
                # place = self.calculate_place(pipeline.group('iop_id'))
                IopId = pipeline.group('iop_id')
                MnemonicKey = pipeline.group('mnemonic_key')
                MnemonicValue = pipeline.group('mnemonic_value')
                InstAddrLow32 = pipeline.group('inst_addr_low32')
                DataAddrLow32 = pipeline.group('data_addr_low32')
                Pipe = pipeline.group('pipe')
                new_inst = instruction(IopId=int(IopId), MnemonicValue=MnemonicValue, InstAddrLow32=InstAddrLow32,
                                       DataAddrLow32=DataAddrLow32, Pipe=Pipe, filename=fname)
                haveSameInst = False
                # print(MnemonicKey)
                # print(IopId,MnemonicValue,InstAddrLow32,DataAddrLow32,Pipe)
                insts = self.instruction_dict.setdefault(MnemonicKey, Instruction())
                for index, inst in enumerate(insts.list):
                    if inst.MnemonicValue == MnemonicValue and inst.InstAddrLow32 == InstAddrLow32 and inst.DataAddrLow32 == DataAddrLow32:
                        if str(inst.IopId) == str(IopId):
                            self.instruction_dict[MnemonicKey].list[index].Pipe += Pipe
                            haveSameInst = True
                            break
                        else:
                            haveSameInst = False
                else:
                    pass

                if not haveSameInst:
                    # print(line)
                    self.instruction_dict.setdefault(MnemonicKey, Instruction()).list.append(new_inst)
                    haveSameInst = False
                # print(line)
            else:
                # print("pipeline is empty")
                pass
        if start:
            return True
        else:
            return False

    def calculate_cycles(self):
        exe_pattern = re.compile(r'.(E+)[^EI]*f\.*C')
        issue_pattern = re.compile(r'M([^uSdsI]*)I[^sS]*f\.*C')
        for this_key, this_insts in self.instruction_dict.items():
            for index, val in enumerate(this_insts.list):
                exe = exe_pattern.search(val.Pipe)
                issue = issue_pattern.search(val.Pipe)
                if exe:
                    # print(exe.group(0))
                    self.instruction_dict[this_key].list[index].exeCycle = ExeCycle(exe.group(0).count('E'), 1, 1)
                if issue:
                    # print(exe.group(0))
                    self.instruction_dict[this_key].list[index].issueCycle = IssueCycle(len(issue.group(1)), 1, 1)

    # def calculate_issue_cycles(self):
    #     issue_pattern = re.compile(r'M([^s]*)I.*f\.*C')
    #     for this_key, this_insts in self.instruction_dict.items():
    #         for index, val in enumerate(this_insts.list):
    #             issue = issue_pattern.search(val.Pipe)
    #             if issue:
    #                 # print(exe.group(0))
    #                 self.instruction_dict[this_key].list[index].issueCycle = IssueCycle(len(exe.group(0)), 1, 1)

    def calculate_count_inst_before_merge(self):
        for key, insts in self.instruction_dict.items():
            self.instruction_dict.setdefault(key, Instruction()).count = len(insts.list)

    def calculate_frequency_inst(self):
        pass

    # 计算每指令每个发射周期的出现的次数
    def issueCycleCount(self):
        new_insts: list[instruction] = []
        new_inst = None
        for key, insts in self.instruction_dict.items():
            new_insts.clear()
            for index, inst in enumerate(insts.list):
                for new_inst_index, new_inst in enumerate(new_insts):
                    if inst.issueCycle.cycle_num == new_inst.issueCycle.cycle_num:
                        self.instruction_dict[key].issueCycle.setdefault(inst.issueCycle.cycle_num,IssueCycle(inst.issueCycle.cycle_num)).cycle_count += 1
                        break
                    elif inst.issueCycle.cycle_num == -1:
                        break
                else:
                    if inst.issueCycle.cycle_num != -1:
                        self.instruction_dict[key].issueCycle.setdefault(inst.issueCycle.cycle_num,IssueCycle(inst.issueCycle.cycle_num)).cycle_count = 1
                        new_insts.append(inst)

    def exeCycleCount(self):
        new_insts: list[instruction] = []
        new_inst = None
        for key, insts in self.instruction_dict.items():
            new_insts.clear()
            for index, inst in enumerate(insts.list):
                for new_inst_index, new_inst in enumerate(new_insts):
                    if inst.exeCycle.cycle_num == new_inst.exeCycle.cycle_num:
                        self.instruction_dict[key].exeCycle.setdefault(inst.exeCycle.cycle_num,ExeCycle(inst.exeCycle.cycle_num)).cycle_count += 1
                        break
                    elif inst.exeCycle.cycle_num == -1:
                        break
                else:
                    if inst.exeCycle.cycle_num != -1:
                        self.instruction_dict[key].exeCycle.setdefault(inst.exeCycle.cycle_num,ExeCycle(inst.exeCycle.cycle_num)).cycle_count = 1
                        new_insts.append(inst)

    # 精简的流水线图,去除每种指令中执行周期数相同的出现，并计算每指令每个执行周期的出现的次数
    def tidy(self):
        new_insts: list[instruction] = []
        new_inst = None
        for key, insts in self.instruction_dict.items():
            new_insts.clear()
            for index, inst in enumerate(insts.list):
                for new_inst_index, new_inst in enumerate(new_insts):
                    if inst.exeCycle.cycle_num == new_inst.exeCycle.cycle_num:
                        new_insts[new_inst_index].exeCycle.cycle_count += 1
                        break
                    elif inst.exeCycle.cycle_num == -1:
                        break
                else:
                    if inst.exeCycle.cycle_num != -1:
                        new_insts.append(inst)
            self.instruction_dict[key].list.clear()
            self.instruction_dict[key].list.extend(new_insts)

    # 还原流水线图
    def calculate_place(self, iop_id):
        pass

    # 写入文本
    def wirte(self, write_path="../data/pipeline_graph/", write_name="", file_type="txt"):
        this_fname = write_name
        if write_name == "":
            this_fname = self.fname
        tup_path = os.path.splitext(this_fname)
        if tup_path[1] == ".txt":
            this_fname = tup_path[0] + "." + file_type
        if "." not in this_fname:
            this_fname += "." + file_type
        # this_fname = self.fname.removesuffix(".txt") + "."+file_type
        # print("basename: "+basename(this_fname))
        with open(write_path + basename(this_fname), 'w', encoding='utf-8') as f:
            f.write(format("MNEMONIC-L", "<20") +
                    format("IOP-ID", "<20") +
                    format("Benchmark", "<20") +
                    # format("Location", "<20") +
                    format("Pipeline-File", "<60") +
                    format("MNEMONIC-R", "<20") +
                    format("EXE-CYCLES", "<20") +
                    format("ISSUE-CYCLES", "<20") +
                    format("INST-ADDR-LOW-32", "<20") +
                    format("DATA-ADDR-LOW-32", "<20") + "Pipeline-Graph" +
                    '\n')
            pipeline_file_pattern = re.compile(
                r'(?P<begin>\d+)_(?P<end>\d+)_(?P<simpt>\d+)_(?P<interval>\d+)_(?P<benchmark>.*)\.txt')
            for this_key, this_insts in self.instruction_dict.items():
                # f.write(this_key + '\n')
                for i in this_insts.list:
                    pipeline_file_name_group = pipeline_file_pattern.search(os.path.basename(i.filename))
                    if pipeline_file_name_group:
                        if len(pipeline_file_name_group.groups()) == 5:
                            i.set_benchmark(pipeline_file_name_group.group("benchmark"))
                            i.set_location(
                                int(pipeline_file_name_group.group("simpt"))*int(pipeline_file_name_group.group("interval"))+i.IopId
                            )
                    f.write(format(this_key, "<20") +
                            format(str(i.IopId), "<20") +
                            format(str(i.benchmark), "<20") +
                            # format(str(i.location), "<20") +
                            format(os.path.basename(str(i.filename)), "<60") +
                            format(i.MnemonicValue, "<20") +
                            format(str(i.exeCycle.cycle_num), "<20") +
                            format(str(i.issueCycle.cycle_num), "<20") +
                            format(i.InstAddrLow32, "<20") +
                            format(i.DataAddrLow32, "<20") + i.Pipe +
                            '\n')

    # 多线程
    def run(self):
        tup_path = os.path.splitext(basename(self.fname))
        for f in self.existed_files:
            if tup_path[0] in f:
                # print("EXISTED: " + self.fname)
                return {"name": self.fname, "state": "EXISTED", "thread": self}
        if self.m1pipe_grep(self.fname):
            self.calculate_cycles()
            self.calculate_count_inst_before_merge()
            self.wirte()
            self.issueCycleCount()
            self.exeCycleCount()
            self.tidy()
            # sleep(10)
            # print(self.instruction_dict.setdefault("addi", Instruction()).list[0].MnemonicValue)
            # print(self.fname + " done")
            return {"name": self.fname, "state": "RE-CONSTRUCTED", "thread": self}
        else:
            return {"name": self.fname, "state": "NOT-PIPELINE_GRAPH", "thread": self}

    # exe_cycle编码
    def marshal(self, exe_cycle_list):
        s = ""
        for e in exe_cycle_list:
            s += str(e.cycle_num) + "(" + str(e.cycle_frequency) + "," + str(e.cycle_count) + ")"
        return s

    # exe_cycle解码
    def unmarshal(self, exe_cycles):
        exe_cycle_list = []
        exe_cycles_iter = re.finditer(
            r'(?P<cycle_num>\d+)\((?P<cycle_frequency>0\.\d+),(?P<cycle_count>\d+)\)', exe_cycles
        )
        for exe_cycle in exe_cycles_iter:
            e = ExeCycle(
                exe_cycle.group("cycle_num"),
                exe_cycle.group("cycle_frequency"),
                exe_cycle.group("cycle_count")
            )
            exe_cycle_list.append(e)
        return exe_cycle_list

    # 根据exe_cycles_old 如 2(0.3,10),3(0.7,5) ,计算出添加exe_cycle_new_num后的周期数和频率
    def merge_exe_cycle(self, exe_cycle_list_old, exe_cycle_list_new):
        merge = []
        for e_old in exe_cycle_list_old:
            for e_new in exe_cycle_list_new:
                pass

    # 合并后使用
    def calculate_exe_cycle_frequency(self):
        new_insts: list[Instruction] = []
        new_inst = None
        cycle_count_sum = 0
        for key, insts in self.instruction_dict.items():
            cycle_count_sum = 0
            for CycleNum, CycleObj in insts.exeCycle.items():
                cycle_count_sum += CycleObj.cycle_count
            for CycleNum, CycleObj in insts.exeCycle.items():
                self.instruction_dict.setdefault(key,Instruction()).exeCycle.setdefault(CycleNum,ExeCycle(CycleNum)).cycle_frequency = \
                    CycleObj.cycle_count/cycle_count_sum

    def calculate_issue_cycle_frequency(self):
        new_insts: list[Instruction] = []
        new_inst = None
        issue_cycle_count_sum = 0
        for key, insts in self.instruction_dict.items():
            issue_cycle_count_sum = 0
            for issueCycleNum, issueCycleObj in insts.issueCycle.items():
                issue_cycle_count_sum += issueCycleObj.cycle_count
            for issueCycleNum, issueCycleObj in insts.issueCycle.items():
                self.instruction_dict.setdefault(key,Instruction()).issueCycle.setdefault(issueCycleNum,IssueCycle(issueCycleNum)).cycle_frequency =\
                    issueCycleObj.cycle_count/issue_cycle_count_sum
    def list2str_sort(self,cycle_list):
        exe_cycles = ""
        if len(cycle_list) == 1:
            exe_cycles = str(cycle_list[0].cycle_num) + "(" + \
                         str(cycle_list[0].cycle_frequency) + "," + \
                         str(cycle_list[0].cycle_count) + "/" + str(
                cycle_list[0].cycle_count) + ")"
        elif len(cycle_list) > 1:
            # print(new_insts)
            exe_cycles_list = [[]]
            exe_cycles_list.clear()
            for inst in cycle_list:
                exe_cycles_list.append(
                    [
                        inst.cycle_num,
                        inst.cycle_frequency,
                        inst.cycle_count,
                        int(inst.cycle_count / inst.cycle_frequency)
                    ]
                )
            exe_cycles_list.sort(key=lambda x: (-x[1], -x[2]))
            for e in exe_cycles_list:
                # print(e)
                exe_cycles += (
                        str(e[0]) + "(" +
                        format(e[1], ".3f") + "," +
                        str(e[2]) + "/" + str(e[3]) +
                        ")" + ";"
                )
        else:
            # continue
            exe_cycles = ""
        return exe_cycles

    def sort(self, source_csv_file, write_path="../data/pipeline_result/", write_name="P8_Insts.csv"):
        p8_insts_headers = ["MNEMONIC", "FREQUENCY", "EXE_CYCLES", "ISSUE_CYCLES", "STATUS", "CATEGORY", "VERSION", "PDF_REAL", "DESC"]
        exe_cycle_set = set()
        exe_cycle = ""
        # if os.path.isfile("../data/p8_insts.csv"):
        #     p8_insts_class = namedtuple('p8_insts_class', p8_insts_headers)
        #     with open("../data/p8_insts.csv", "r", encoding='utf-8') as fr:
        #         with open("../data/_p8_insts.csv", "w+", encoding='utf-8') as fw:
        #             csv_reader = csv.reader(fr)
        #             _p8_insts_csv = csv.writer(fw)
        #             for r in csv_reader:
        #                 row_info = p8_insts_class(*r)
        #                 new_insts = self.instruction.setdefault(row_info.MNEMONIC, [])
        #                 if len(new_insts) == 1:
        #                     row_info.EXE_CYCLES += str(new_insts[0].exeCycle.cycle_num)
        #                 elif len(new_insts) > 1:
        #                     for inst in self.instruction[row_info.MNEMONIC]:
        #                         row_info.EXE_CYCLES += str(inst.EXE_Cycles) + ";"
        #                 else:
        #                     pass
        #                 # _p8_insts_csv.writerow(
        #                 #     [row_info.MNEMONIC, row_info.EXE_CYCLES, row_info.STATUS, row_info.CATEGORY,
        #                 #      row_info.VERSION, row_info.PDF_REAL, row_info.DESC])
        # else:
        with open(source_csv_file, "r", encoding='utf-8') as fr:
            csv_reader = csv.reader(fr)
            # headers=next(csv_reader)
            headers = ["MNEMONIC_L0", "MNEMONIC_L1", "MNEMONIC_L2", "STATUS", "VERSION", "CATEGORY", "PDF_SHOW",
                       "PDF_REAL", "DESC"]
            p8_insts_class = namedtuple('p8_insts_class', headers)
            with open(write_path + write_name, 'w+', encoding="utf-8") as fw:
                p8_insts_csv = csv.writer(fw)
                p8_insts_csv.writerow(p8_insts_headers)
                for r in csv_reader:
                    row_info = p8_insts_class(*r)
                    # print(row_info)
                    exe_cycles = ""
                    insts = self.instruction_dict.setdefault(row_info.MNEMONIC_L2, Instruction())
                    exe_cycles = self.list2str_sort(list(insts.exeCycle.values()))
                    issue_insts = list(self.instruction_dict.setdefault(row_info.MNEMONIC_L2, Instruction()).issueCycle.values())
                    issue_cycles = self.list2str_sort(issue_insts)
                    p8_insts_csv.writerow(
                        [
                            row_info.MNEMONIC_L2,
                            "(" + format(insts.frequency, ".3f") + "," + str(insts.count) + "/" + str(self.inst_count) + ")",
                            exe_cycles, issue_cycles, row_info.STATUS, row_info.CATEGORY,
                            row_info.VERSION, row_info.PDF_REAL, row_info.DESC
                        ]
                    )

    @staticmethod
    def get_existed():
        existed_files = []
        for root, dirs, files in os.walk("../data/pipeline_graph/", topdown=False):
            for name in files:
                existed_files.append(os.path.join(root, name))
        return existed_files

# 进行对比, 流水线图中抓取的执行周期数目跟UM10.16表格
def insertUMinfo(source_csv_file="../data/20220829-P8_Insts.csv",write_path="../data/",write_name="20220908_p8_instr.csv"):
    # 读取UM中整理好的数据
    workbook = load_workbook(filename="../data/meta/UM_10_16Table.xlsx")
    sheet1 = workbook.active
    with open(source_csv_file, "r", encoding='utf-8') as fr:
        csv_reader = csv.reader(fr)
        next(csv_reader)
        headers = ["MNEMONIC","FREQUENCY","EXE_CYCLES","ISSUE_CYCLES","CATEGORY","VERSION","PDF_REAL","DESC"]
        p8_insts_headers = ["MNEMONIC","FREQUENCY","EXE_CYCLES","UM_DESC","ISSUE_CYCLES","CATEGORY","VERSION","PDF_REAL","DESC"]
        p8_insts_class = namedtuple('p8_insts_class', headers)
        with open(write_path + write_name, 'w+', encoding="utf-8") as fw:
            p8_insts_csv = csv.writer(fw)
            p8_insts_csv.writerow(p8_insts_headers)
            UM_MNEMONIC_dict = {}
            # 从UM里查询
            for row_num, row_cells in enumerate(sheet1["1:" + str(sheet1.max_row)], start=1):
                MNEMONICS = str(row_cells[0].value).split(" ")
                for MNEMONIC in MNEMONICS:
                    value = str(row_cells[2].value)
                    if value == "None":
                        value=""
                    UM_MNEMONIC_dict.setdefault(MNEMONIC, value)
            for r in csv_reader:
                row_info = p8_insts_class(*r)
                print(row_info.MNEMONIC)
                if UM_MNEMONIC_dict.setdefault(row_info.MNEMONIC, "")!=UM_MNEMONIC_dict.setdefault(row_info.MNEMONIC+".", "") and UM_MNEMONIC_dict.setdefault(row_info.MNEMONIC+".", "") != "":
                    if UM_MNEMONIC_dict.setdefault(row_info.MNEMONIC, "")!="":
                        UM_DESC=UM_MNEMONIC_dict.setdefault(row_info.MNEMONIC, "")+"; if with dot suffix: "+UM_MNEMONIC_dict.setdefault(row_info.MNEMONIC+".", "")
                    else:
                        UM_DESC="only with dot suffix: "+UM_MNEMONIC_dict.setdefault(row_info.MNEMONIC+".", "")
                else:
                    UM_DESC = UM_MNEMONIC_dict.setdefault(row_info.MNEMONIC, "")
                # 写入
                p8_insts_csv.writerow(
                    [
                        row_info.MNEMONIC,row_info.FREQUENCY,row_info.EXE_CYCLES,
                        UM_DESC,
                        row_info.ISSUE_CYCLES,row_info.CATEGORY,row_info.VERSION,row_info.PDF_REAL,row_info.DESC
                    ]
                )
# def t():
#     a = [1, 2, 3, 4, 5, 6]
#     i = 0
#     for i in a:
#         if i == 6: break
#     else:
#         print("end:" + str(i))

return_TRACE = []
threadLock = threading.Lock()


def job_done(this_future: Future):
    # sleep(5)
    job = this_future.result()
    threadLock.acquire()
    return_TRACE.append(job["thread"])
    threadLock.release()
    print(job["state"] + " : " + job["name"])


def pd():
    pass


if __name__ == '__main__':
    threads = []
    start_time = datetime.datetime.now()
    # 0. 检查
    existed_files = Trace.get_existed()
    # 1. 并行
    # pipe_list_temp = runcmd(["find ../*.txt"])
    # pipe_list_temp = ["../500001_505000_1_5000000_523.xalancbmk_r.txt"]

    pipe_list_temp = runcmd(["find ../runspec_gem5_power/*r/pipe_result/ -name '*.txt'"])
    runcmd(["mkdir -p ../data/pipeline_graph/"])
    runcmd(["mkdir -p ../data/pipeline_result/"])
    pipe_list = list(sorted(set(pipe_list_temp)))
    # [print(f) for f in pipe_list]
    if "" in pipe_list:
        pipe_list.remove("")
    # [print(f) for f in pipe_list]

    for p in pipe_list:
        t = Trace(fname=p, existed_files=existed_files)
        threads.append(t)

    # 线程异步回调
    threads_num = 20
    process_pool = ProcessPoolExecutor(max_workers=threads_num)
    todo = []
    for index, t in enumerate(threads):
        future = process_pool.submit(threads[index].run)
        todo.append(future)
        future.add_done_callback(job_done)
    process_pool.shutdown(wait=True)
    for t in return_TRACE:
        pass

    # 2. 合并(按指令)
    merge = Trace(fname="merge", existed_files=existed_files)
    for t in return_TRACE:
        inst_count = 0
        for key, insts in t.instruction_dict.items():
            merge.inst_count += insts.count
            merge.instruction_dict.setdefault(key, Instruction()).count += insts.count
            merge.instruction_dict.setdefault(key, Instruction()).list.extend(insts.list)

            # 统计相同助记符下出现的不同执行周期数
            for exeCycleNum,exeCycleObj in insts.exeCycle.items():
                merge.instruction_dict.setdefault(key,Instruction()).exeCycle.setdefault(exeCycleNum,ExeCycle(exeCycleNum)).cycle_count += exeCycleObj.cycle_count
            # 统计相同助记符下出现的不同发射周期数
            for issueCycleNum,issueCycleObj in insts.issueCycle.items():
                merge.instruction_dict.setdefault(key,Instruction()).issueCycle.setdefault(issueCycleNum,IssueCycle(issueCycleNum)).cycle_count += issueCycleObj.cycle_count
            # # print(key)
    # 计算指令频率
    for key, insts in merge.instruction_dict.items():
        merge.instruction_dict.setdefault(key, Instruction()).frequency = \
            merge.instruction_dict.setdefault(key, Instruction()).count / merge.inst_count
    merge.wirte(write_path="../data/pipeline_result/", write_name="mid")
    merge.tidy()
    print("TIDY-ENDED")
    merge.wirte(write_path="../data/pipeline_result/", write_name=start_time.strftime('%Y%m%d') + "-" + merge.fname)
    print("MERGE-ENDED")
    merge.calculate_issue_cycle_frequency()
    merge.calculate_exe_cycle_frequency()
    merge.sort(source_csv_file="../data/meta/scripts_csv_power-isa-implementation.csv",
               write_path="../data/pipeline_result/", write_name=start_time.strftime('%Y%m%d') + "-" + "P8_Insts.csv")
    print("SORT-ENDED")
    end_time = datetime.datetime.now()
    # runcmd("cp -r"+start_time.strftime('%Y%m%d%H%M%S'))
    print("CONSUMED TIME", end_time - start_time)
