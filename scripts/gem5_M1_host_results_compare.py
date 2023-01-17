import csv
import re
import subprocess
import datetime
import sys
import os
from collections import defaultdict
from collections import namedtuple
from typing import Dict, Any

import openpyxl
from pandas import read_csv
from openpyxl import load_workbook, Workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# 在gem5spec目录下调用
# 改
begin_time=str(sys.argv[1])
# begin_time="20220915"
gem5_ckp_results_csv = "./data/gem5/"+begin_time+"/Each_case_ckp_data.csv"
M1_ckp_results_csv = "./data/M1/each_bm_cpt_m1.csv"
# 改
gen_file="./data/gem5/"+begin_time+"-comparison_M1_gem5_SPEC2017_sampling_results" + ".xlsx"
# gen_file="../data/gem5/"+begin_time+"_comparison_M1_gem5_SPEC2017_sampling_results" + ".xlsx"

class QtraceSkip:
    def __init__(self,expectSkipVgi,realSkipVgi,Valid):
        self.expectSkipVgi=expectSkipVgi
        self.realSkipVgi=realSkipVgi
        self.valid=Valid

# qtskip[500.perlbench_r][simpts].
qtskip: Dict[str, Dict[str, QtraceSkip]] = {}


def runcmd(command):
    ret = subprocess.run(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, encoding="utf-8",
                         timeout=None)
    if ret.returncode == 0:
        # print("success:", ret)
        return ret.stdout.split('\n')
    else:
        # print("error:", ret)
        return ret.stderr.split('\n')


# 自动设置列宽
def auto_set_column_width(worksheet):
    dims = {}
    for row in worksheet.rows:
        for cell in row:
            # 跳过公式
            if cell.value:
                if str(cell.value)[0]=="=":
                    # print(cell.value)
                    continue
                # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                # 大写字母的长度近似多0.45
                cell_len = 0.7 * len(re.findall('([\u4E00-\u9FA5])', str(cell.value))) + \
                           len(str(cell.value)) + 0.45 * len(re.findall('([A-Z])', str(cell.value)))
                # print(float(str(cell.font.size)))
                # print(cell.column)
                if cell.font.size:
                    if float(cell.font.size) > 11:
                        cell_len += float(cell.font.size) - 11
                if cell.font.bold:
                    cell_len += 1
                dims[cell.column] = max((dims.setdefault(cell.column, 0), cell_len))
    for col, value in dims.items():
        # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+2是用来调整最终效果的
        worksheet.column_dimensions[get_column_letter(col)].width = value + 1
        # print(col,value)


# 改变字体的样式,大小,粗体,斜体;wss表示[sheet..]
def pyCellFontStyle(wss, onlyRow=False, rowNumStart: int = 1, onlyCol=False,
                    colNumStart: int = 1, rowAndCol=False, fontName="Times New Roman", fontSize=11, isbold=False,
                    italic=False,set_align=False,horizontal='center',vertical='center',set_border=False,border_style="thin",border_color="000000",
                    border_bottom=False):
    # border_style {'hair', 'mediumDashed', 'thin', 'medium', 'double', 'dashDot', 'mediumDashDotDot', 'mediumDashDot', 'dashDotDot', 'dashed', 'dotted', 'thick', 'slantDashDot'}
    side = Side(border_style=border_style, color=border_color)
    if onlyRow:
        for ws in wss:
            nrows = ws.max_row  # 获得行数
            ncols = ws.max_column
            for j in range(ncols):
                if j + 1 > ncols:
                    return
                # name代表样式，size代表大小，bold代表粗体，italic代表斜体
                ws.cell(rowNumStart, j + 1).font = Font(name=fontName, size=fontSize, bold=isbold, italic=italic)
                if set_align:
                    ws.cell(rowNumStart,j + 1).alignment=Alignment(horizontal=horizontal, vertical=vertical)
                if set_border:
                    ws.cell(rowNumStart,j + 1).border = Border(top=side,bottom=side,left=side,right=side)
                    if border_bottom:
                        ws.cell(rowNumStart,j + 1).border = Border(bottom=side)
    if onlyCol:
        for ws in wss:
            nrows = ws.max_row  # 获得行数
            ncols = ws.max_column
            for i in range(nrows):
                if i + 1> nrows:
                    return
                # name代表样式，size代表大小，bold代表粗体，italic代表斜体
                ws.cell(i + 1, colNumStart).font = Font(name=fontName, size=fontSize, bold=isbold, italic=italic)
                if set_align:
                    ws.cell(i + 1, colNumStart).alignment=Alignment(horizontal=horizontal, vertical=vertical)
                if set_border:
                    ws.cell(i + 1, colNumStart).border = Border(top=side,bottom=side,left=side,right=side)
                    if border_bottom:
                        ws.cell(i + 1, colNumStart).border = Border(bottom=side)
    if rowAndCol:
            for ws in wss:
                nrows = ws.max_row  # 获得行数
                ncols = ws.max_column
                for i in range(nrows):
                    if rowNumStart + i>nrows:
                        continue
                    for j in range(ncols):
                        # name代表样式，size代表大小，bold代表粗体，italic代表斜体
                        if colNumStart + j>ncols:
                            continue
                        ws.cell(rowNumStart + i, colNumStart + j).font = Font(name=fontName, size=fontSize,bold=isbold, italic=italic)
                        if set_align:
                            ws.cell(rowNumStart + i, colNumStart + j).alignment=Alignment(horizontal=horizontal, vertical=vertical)
                        if set_border:
                            ws.cell(rowNumStart + i, colNumStart + j).border = Border(top=side,bottom=side,left=side,right=side)
                            if border_bottom:
                                ws.cell(rowNumStart + i, colNumStart + j).border = Border(bottom=side)
                        # print(rowNumStart + i,colNumStart + j,ws.cell(rowNumStart + i, colNumStart + j).value)

# copy sheet from src->dst, 当默认Sheet1时，默认直接写入，不用指定sheet name
def copy_sheet(src_xlsx,dst_xlsx,src_sheetname="Sheet1",dst_sheetname="Sheet1"):
    sw=load_workbook(f'{src_xlsx}')
    dw=load_workbook(f'{dst_xlsx}')
    src_sheet=sw[f'{src_sheetname}']
    dst_sheet = dw[f'{dst_sheetname}']

    for row in src_sheet.iter_rows():
        row_list=[]
        for cell in row:
            row_list.append(cell.value)
        dst_sheet.append(row_list)
    dw.save(f'{dst_xlsx}')


def read_qt_skip(qtskip_excel_file="./data/meta/20221214-qt-skip.xlsx"):
    workbook = load_workbook(filename=qtskip_excel_file, data_only=True)
    sheet=workbook.active
    for row_num,row_cells in enumerate(sheet["2:"+str(sheet.max_row)],start=2):
        qtskip.setdefault(row_cells[0].value,{}).setdefault(
            str(row_cells[1].value),
            QtraceSkip(
                str(row_cells[2].value),
                str(row_cells[3].value),
                str(row_cells[4].value))
        )
    # print(qtskip.get("500.perlbench_r",{}).get("359",QtraceSkip(None,None,None)).realSkipVgi)
    pass

def grep_cache_data_gem5(benchmark="",simpt="",num="",gem5spec_path="."):
    L1Icache_hits, L1Icache_misses, L1Icache_accesses = '', '', ''
    L1Dcache_hits, L1Dcache_misses, L1Dcache_accesses = '', '', ''
    L2cache_hits, L2cache_misses, L2cache_accesses = '', '', ''
    L3cache_hits, L3cache_misses, L3cache_accesses = '', '', ''
    validstats=0
    L1Icache_miss_rate,L1Icache_miss,L1Icache_access,L1Icache_MPKI, \
    L1Dcache_miss_rate,L1Dcache_miss,L1Dcache_access,L1Dcache_MPKI, \
    L2cache_miss_rate,L2cache_miss,L2cache_access,L2cache_MPKI, \
    L3cache_miss_rate,L3cache_miss,L3cache_access,L3cache_MPKI= \
    None, None, None, None,None,None,None,None,None,None,None,None,None,None,None,None

    try:
        with open(f"{gem5spec_path}/runspec_gem5_power/{benchmark}/output_ckp{num}/stats.txt", 'r', encoding='utf-8') as f:
            for line in f.readlines():

                if "Begin Simulation Statistics" in line:
                    validstats+=1

                if validstats==2:
                    icache_hits = re.search(
                        r'(?P<itemName>system.ruby.l1_cntrl0.L1Icache.m_demand_hits)\s+(?P<itemValue>((\-|\+)?\d+(\.\d+)?))\s+(?P<itemComments>#.*)',
                        line)
                    if icache_hits:
                        L1Icache_hits = icache_hits.group("itemValue")
                        continue
                    icache_misses = re.search(
                        r'(?P<itemName>system.ruby.l1_cntrl0.L1Icache.m_demand_misses)\s+(?P<itemValue>((\-|\+)?\d+(\.\d+)?))\s+(?P<itemComments>#.*)',
                        line)
                    if icache_misses:
                        L1Icache_misses = icache_misses.group("itemValue")
                        continue
                    icache_accesses = re.search(
                        r'(?P<itemName>system.ruby.l1_cntrl0.L1Icache.m_demand_accesses)\s+(?P<itemValue>((\-|\+)?\d+(\.\d+)?))\s+(?P<itemComments>#.*)',
                        line)
                    if icache_accesses:
                        L1Icache_accesses = icache_accesses.group("itemValue")
                        continue

                    # L1D
                    dcache_hits = re.search(
                        r'(?P<itemName>system.ruby.l1_cntrl0.L1Dcache.m_demand_hits)\s+(?P<itemValue>((\-|\+)?\d+(\.\d+)?))\s+(?P<itemComments>#.*)',
                        line)
                    if dcache_hits:
                        L1Dcache_hits = dcache_hits.group("itemValue")
                        continue

                    dcache_misses = re.search(
                        r'(?P<itemName>system.ruby.l1_cntrl0.L1Dcache.m_demand_misses)\s+(?P<itemValue>((\-|\+)?\d+(\.\d+)?))\s+(?P<itemComments>#.*)',
                        line)
                    if dcache_misses:
                        L1Dcache_misses = dcache_misses.group("itemValue")
                        continue

                    dcache_accesses = re.search(
                        r'(?P<itemName>system.ruby.l1_cntrl0.L1Dcache.m_demand_accesses)\s+(?P<itemValue>((\-|\+)?\d+(\.\d+)?))\s+(?P<itemComments>#.*)',
                        line)
                    if dcache_accesses:
                        L1Dcache_accesses = dcache_accesses.group("itemValue")
                        continue

                    # L2
                    l2cache_hits = re.search(
                        r'(?P<itemName>system.ruby.l2_cntrl0.L2cache.m_demand_hits)\s+(?P<itemValue>((\-|\+)?\d+(\.\d+)?))\s+(?P<itemComments>#.*)',
                        line)
                    if l2cache_hits:
                        L2cache_hits = l2cache_hits.group("itemValue")
                        continue

                    l2cache_misses = re.search(
                        r'(?P<itemName>system.ruby.l2_cntrl0.L2cache.m_demand_misses)\s+(?P<itemValue>((\-|\+)?\d+(\.\d+)?))\s+(?P<itemComments>#.*)',
                        line)
                    if l2cache_misses:
                        L2cache_misses = l2cache_misses.group("itemValue")
                        continue

                    l2cache_accesses = re.search(
                        r'(?P<itemName>system.ruby.l2_cntrl0.L2cache.m_demand_accesses)\s+(?P<itemValue>((\-|\+)?\d+(\.\d+)?))\s+(?P<itemComments>#.*)',
                        line)
                    if l2cache_accesses:
                        L2cache_accesses = l2cache_accesses.group("itemValue")
                        continue

                        # L3
                    l3cache_hits = re.search(
                        r'(?P<itemName>system.ruby.l3_cntrl0.L3cache.m_demand_hits)\s+(?P<itemValue>((\-|\+)?\d+(\.\d+)?))\s+(?P<itemComments>#.*)',
                        line)
                    if l3cache_hits:
                        L3cache_hits = l3cache_hits.group("itemValue")
                        continue

                    l3cache_misses = re.search(
                        r'(?P<itemName>system.ruby.l3_cntrl0.L3cache.m_demand_misses)\s+(?P<itemValue>((\-|\+)?\d+(\.\d+)?))\s+(?P<itemComments>#.*)',
                        line)
                    if l3cache_misses:
                        L3cache_misses = l3cache_misses.group("itemValue")
                        continue

                    l3cache_accesses = re.search(
                        r'(?P<itemName>system.ruby.l3_cntrl0.L3cache.m_demand_accesses)\s+(?P<itemValue>((\-|\+)?\d+(\.\d+)?))\s+(?P<itemComments>#.*)',
                        line)
                    if l3cache_accesses:
                        L3cache_accesses = l3cache_accesses.group("itemValue")
                        break
    except IOError:
        # print(f"gem5 file exception: {gem5spec_path}/runspec_gem5_power/{benchmark}/output_ckp{num}/stats.txt")
        pass
    finally:

        # print(
        #     "gem5:",
        #     L1Icache_hits, L1Icache_misses, L1Icache_accesses ,
        #     L1Dcache_hits, L1Dcache_misses, L1Dcache_accesses,
        #     L2cache_hits, L2cache_misses, L2cache_accesses,
        #     L3cache_hits, L3cache_misses, L3cache_accesses
        # )
        if L1Icache_hits!='' and L1Icache_misses!='' and L1Icache_accesses!='':
            L1Icache_miss_rate=float(L1Icache_misses)/(float(L1Icache_accesses))
            L1Icache_miss=float(L1Icache_misses)
            L1Icache_access=float(L1Icache_accesses)
            L1Icache_MPKI=1000*float(L1Icache_miss)/5000000

        if L1Dcache_hits!='' and L1Dcache_misses!='' and L1Dcache_accesses!='':
            L1Dcache_miss_rate=float(L1Dcache_misses)/(float(L1Dcache_accesses))
            L1Dcache_miss=float(L1Dcache_misses)
            L1Dcache_access=float(L1Dcache_accesses)
            L1Dcache_MPKI=1000*float(L1Dcache_miss)/5000000

        if L2cache_hits!='' and L2cache_misses!='' and L2cache_accesses!='':
            L2cache_miss_rate=float(L2cache_misses)/(float(L2cache_accesses))
            L2cache_miss=float(L2cache_misses)
            L2cache_access=float(L2cache_accesses)
            L2cache_MPKI=1000*float(L2cache_miss)/5000000

        if L3cache_hits!='' and L3cache_misses!='' and L3cache_accesses!='':
            L3cache_miss_rate=float(L3cache_misses)/(float(L3cache_accesses))
            L3cache_miss=float(L3cache_misses)
            L3cache_access=float(L3cache_accesses)
            L3cache_MPKI=1000*float(L3cache_miss)/5000000

        # print("gem5:",L1Icache_miss_rate,L1Dcache_miss_rate,L2cache_miss_rate,L3cache_miss_rate)
        return \
            L1Icache_miss_rate,L1Icache_miss,L1Icache_access,L1Icache_MPKI, \
            L1Dcache_miss_rate,L1Dcache_miss,L1Dcache_access,L1Dcache_MPKI, \
            L2cache_miss_rate,L2cache_miss,L2cache_access,L2cache_MPKI, \
            L3cache_miss_rate,L3cache_miss,L3cache_access,L3cache_MPKI

def grep_cache_data_M1(benchmark="",simpt="",gem5spec_M1_v0_path="./data/M1/gem5spec_v0_M1"):
    validstats=0

    L2cache_loads, L2cache_stores, L2cache_l1touches, L2cache_l2touches = '', '', '', ''
    L3cache_loads, L3cache_stores, L3cache_l1touches, L3cache_l2touches = '', '', '', ''
    L3_1cache_loads, L3_1cache_stores, L3_1cache_l1touches, L3_1cache_l2touches = '', '', '', ''
    Memory_cache_loads, Memory_cache_stores, Memory_cache_l1touches, Memory_cache_l2touches = '', '', '', ''

    L1Icache_miss_rate,L1Icache_miss,L1Icache_access,L1Icache_MPKI,\
    L1Dcache_miss_rate,L1Dcache_miss,L1Dcache_access,L1Dcache_MPKI,\
    L2cache_miss_rate,L2cache_miss,L2cache_access,L2cache_MPKI,\
    L3cache_miss_rate,L3cache_miss,L3cache_access,L3cache_MPKI=\
    None, None, None, None,None,None,None,None,None,None,None,None,None,None,None,None

    validstats=0
    try:
        with open(f"{gem5spec_M1_v0_path}/runspec_gem5_power/{benchmark}/M1_result/{simpt}_5000000_{benchmark}.results", 'r', encoding='utf-8') as f:
            for line in f.readlines():

                if "Timer Statistics Generated During Run" in line:
                    validstats+=1

                if validstats==1:

                    # L1I
                    # null

                    # L1D
                    _L1Dcache_miss_rate = re.search(
                        r'(?P<itemName>a load that has never been rejected will hit in L1)\s+=\s+(?P<itemValue>((\-|\+)?\d+(\.\d+)?))\s+(?P<itemComments>.*)',
                        line)
                    if _L1Dcache_miss_rate:
                        L1Dcache_miss_rate = 1-float(_L1Dcache_miss_rate.group("itemValue"))
                        continue

                    # L2
                    l2cache = re.search(
                        r'(?P<itemName>L2)\s+(?P<loads>((\-|\+)?\d+(\.\d+)?))\s+(?P<l1touches>((\-|\+)?\d+(\.\d+)?))\s+(?P<l2touches>((\-|\+)?\d+(\.\d+)?))\s+(?P<stores>((\-|\+)?\d+(\.\d+)?))',
                        line)
                    if l2cache:
                        L2cache_loads = l2cache.group("loads")
                        L2cache_stores = l2cache.group("stores")
                        L2cache_l1touches=l2cache.group("l1touches")
                        L2cache_l2touches=l2cache.group("l2touches")
                        continue

                    # L3
                    l3cache = re.search(
                        r'(?P<itemName>L3)\s+(?P<loads>((\-|\+)?\d+(\.\d+)?))\s+(?P<l1touches>((\-|\+)?\d+(\.\d+)?))\s+(?P<l2touches>((\-|\+)?\d+(\.\d+)?))\s+(?P<stores>((\-|\+)?\d+(\.\d+)?))',
                        line)
                    if l3cache:
                        L3cache_loads = l3cache.group("loads")
                        L3cache_stores = l3cache.group("stores")
                        L3cache_l1touches=l3cache.group("l1touches")
                        L3cache_l2touches=l3cache.group("l2touches")
                        continue

                    # L3.1
                    l3_1cache = re.search(
                        r'(?P<itemName>L3.1)\s+(?P<loads>((\-|\+)?\d+(\.\d+)?))\s+(?P<l1touches>((\-|\+)?\d+(\.\d+)?))\s+(?P<l2touches>((\-|\+)?\d+(\.\d+)?))\s+(?P<stores>((\-|\+)?\d+(\.\d+)?))',
                        line)
                    if l3_1cache:
                        L3_1cache_loads = l3_1cache.group("loads")
                        L3_1cache_stores = l3_1cache.group("stores")
                        L3_1cache_l1touches=l3_1cache.group("l1touches")
                        L3_1cache_l2touches=l3_1cache.group("l2touches")
                        continue

                    # Memory
                    memory_cache = re.search(
                        r'(?P<itemName>Memory)\s+(?P<loads>((\-|\+)?\d+(\.\d+)?))\s+(?P<l1touches>((\-|\+)?\d+(\.\d+)?))\s+(?P<l2touches>((\-|\+)?\d+(\.\d+)?))\s+(?P<stores>((\-|\+)?\d+(\.\d+)?))',
                        line)
                    if memory_cache:
                        Memory_cache_loads = memory_cache.group("loads")
                        Memory_cache_stores = memory_cache.group("stores")
                        Memory_cache_l1touches=memory_cache.group("l1touches")
                        Memory_cache_l2touches=memory_cache.group("l2touches")
                        break
    except IOError:
        # print(f"M1 file Exception: {gem5spec_M1_v0_path}/runspec_gem5_power/{benchmark}/M1_result/{simpt}_5000000_{benchmark}.results")
        pass
    finally:
        # print("M1:",
        #     L2cache_loads, L2cache_stores,L3cache_loads, L3cache_stores,L3_1cache_loads, L3_1cache_stores,Memory_cache_loads, Memory_cache_stores,
        # )
        if L2cache_loads!='' and L2cache_stores!='' and L2cache_l2touches!='' and\
                L3cache_loads!='' and L3cache_stores!='' and L3cache_l2touches!='' and\
                L3_1cache_loads!='' and L3_1cache_stores!='' and L3_1cache_l2touches!='' and\
                Memory_cache_loads!=''and Memory_cache_stores!='' and Memory_cache_l2touches!='':
            L2cache_access=(float(L2cache_loads)+float(L2cache_stores)+float(L2cache_l2touches)+
                            float(L3cache_loads)+float(L3cache_stores)+float(L3cache_l2touches)+
                            float(L3_1cache_loads)+float(L3_1cache_stores)+float(L3_1cache_l2touches)+
                            float(Memory_cache_loads)+float(Memory_cache_stores)+float(Memory_cache_l2touches))
            L2cache_miss=L2cache_access-(float(L2cache_loads)+float(L2cache_stores)+float(L2cache_l2touches))
            L2cache_miss_rate=L2cache_miss/L2cache_access
            L2cache_MPKI=1000*L2cache_miss/5000000

        if L3cache_loads!='' and L3cache_stores!='' and L3cache_l2touches!='' and \
                L3_1cache_loads!='' and L3_1cache_stores!=''and L3_1cache_l2touches!='' and \
                Memory_cache_loads!=''and Memory_cache_stores!=''and Memory_cache_l2touches!='':
            L3cache_access=(float(L3cache_loads)+float(L3cache_stores)+float(L3cache_l2touches)+
                            float(L3_1cache_loads)+float(L3_1cache_stores)+float(L3_1cache_l2touches)+
                            float(Memory_cache_loads)+float(Memory_cache_stores)+float(Memory_cache_l2touches))
            L3cache_miss=L3cache_access-(float(L3cache_loads)+float(L3cache_stores)+float(L3_1cache_loads)+float(L3_1cache_stores)+float(L3cache_l2touches)+float(L3_1cache_l2touches))
            L3cache_miss_rate=L3cache_miss/L3cache_access
            L3cache_MPKI=1000*L3cache_miss/5000000

        # print("M1:",L1Icache_miss_rate,L1Dcache_miss_rate,L2cache_miss_rate,L3cache_miss_rate)
        return \
            L1Icache_miss_rate,L1Icache_miss,L1Icache_access,L1Icache_MPKI, \
            L1Dcache_miss_rate,L1Dcache_miss,L1Dcache_access,L1Dcache_MPKI, \
            L2cache_miss_rate,L2cache_miss,L2cache_access,L2cache_MPKI, \
            L3cache_miss_rate,L3cache_miss,L3cache_access,L3cache_MPKI

def data_pre(M1_source_csv_file=M1_ckp_results_csv, gem5_source_csv_file=gem5_ckp_results_csv,template_excel_path="./data/meta/gem5_statistics_result_template.xlsx"):
    # 改
    runcmd("mkdir -p ./data/gem5/"+begin_time)
    runcmd("paste -d ',' "+M1_source_csv_file+" "+gem5_source_csv_file+" >./data/gem5/M1_gem5_paste.csv")
    excel_path="./temp.xlsx"

    # datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    with open("./data/gem5/M1_gem5_paste.csv", "r", encoding='utf-8') as fr_M1_gem5:
        read_csv(fr_M1_gem5).to_excel(excel_path)

    workbook = load_workbook(filename=excel_path)
    sheet=workbook.active
    sheet.delete_cols(1)
    sheet.insert_cols(2)
    sheet.insert_cols(6)
    sheet.insert_cols(8)
    # print(sheet.max_row)
    sheet.move_range("J1:"+"J"+str(sheet.max_row),rows=0,cols=-8)
    sheet.move_range("M1:"+"M"+str(sheet.max_row),rows=0,cols=-7)
    sheet.move_range("N1:"+"N"+str(sheet.max_row),rows=0,cols=-6)
    sheet.delete_cols(9,4)
    sheet.cell(1,9).value="WeightedCPI Err(Ckp gem5/M1)"
    merge_cols=[]
    delete_rows=[]
    row_num=1
    # 数据清洗
    for row_num,row_cells in enumerate(sheet["2:"+str(sheet.max_row)],start=2):
        if row_cells[0].value == "Benchmark" or row_cells[0].value == None:
            # print(row_num)
            delete_rows.append(row_num)
    for row_num in delete_rows[::-1]:
        sheet.delete_rows(row_num)

    # add qt skip info
    read_qt_skip()

    # print(sheet.max_row)
    # print(sheet.max_column)
    for row_num,row_cells in enumerate(sheet["2:"+str(sheet.max_row)],start=2):
        if qtskip.get(row_cells[0].value,{}).get(str(row_cells[2].value),QtraceSkip(None,None,None)).realSkipVgi!=None:
            sheet.cell(row_num,column_index_from_string("J")).value=int(qtskip.get(row_cells[0].value,{}).get(str(row_cells[2].value),QtraceSkip(None,None,None)).valid)
            sheet.cell(row_num,column_index_from_string("K")).value=int(qtskip.get(row_cells[0].value,{}).get(str(row_cells[2].value),QtraceSkip(None,None,None)).expectSkipVgi)
            sheet.cell(row_num,column_index_from_string("L")).value=int(qtskip.get(row_cells[0].value,{}).get(str(row_cells[2].value),QtraceSkip(None,None,None)).realSkipVgi)
    workbook.save(excel_path)

    workbook2 = load_workbook(filename=template_excel_path)
    sheet0=workbook2.active
    sheet1_CPI=workbook2.create_sheet("summaryCPI",1)
    sheet2_CKPs=workbook2.create_sheet("CKPs",3)
    sheet3=workbook2.create_sheet("summaryCMR",2)
    workbook2.save(gen_file)

    copy_sheet("./temp.xlsx",gen_file,"Sheet1","CKPs")
    runcmd('rm -rf ./temp.xlsx')

def gen_cmp_results(write_path="",write_name="",template_excel_path="./data/meta/gem5_statistics_result_template.xlsx"):
    M1_gem5_ckp_class = namedtuple('M1_gem5_ckp_class',
                                   [
                                       "Benchmark","Checkpoint","Simpts","Weights","CPI_M1","CPI_gem5","WeightedCPI_M1","WeightedCPI_gem5",
                                   ])

    workbook = load_workbook(filename=gen_file)
    sheet0=workbook.active
    sheet1_CPI=workbook["summaryCPI"]
    sheet2_CKPs=workbook["CKPs"]
    sheet3_cache=workbook["summaryCMR"]
    # 处理sheet0

    # 处理sheet1_CPI
    initial_data=[["","Benchmark","Sum WeightedCPI(Ckp M1)","Sum WeightedCPI(Ckp gem5)","Total WeightedCPI(Ckp gem5)","Sum WeightedCPI Err(Ckp gem5/M1)","Credibility(Ckp M1)"],
                  ["int","500.perlbench_r"],["int","502.gcc_r"],["int","505.mcf_r"],["int","520.omnetpp_r"],["int","523.xalancbmk_r"],["int","525.x264_r"],["int","531.deepsjeng_r"],["int","541.leela_r"],["int","548.exchange2_r"],["int","557.xz_r"],
                  ["fp","503.bwaves_r"],["fp","507.cactuBSSN_r"],["fp","508.namd_r"],["fp","510.parest_r"],["fp","511.povray_r"],["fp","519.lbm_r"],
                  ["fp","521.wrf_r"],["fp","526.blender_r"],["fp","527.cam4_r"],["fp","538.imagick_r"],["fp","544.nab_r"],["fp","549.fotonik3d_r"],["fp","554.roms_r"],["","999.specrand_ir"]]
    for info in initial_data:
        sheet1_CPI.append(info)

    initial_data=[["","Benchmark",
                   "Sum Weighted L1I CMR(Ckp M1)","Sum Weighted L1I CMR(Ckp gem5)","Total Weighted L1I CMR(Ckp gem5)",
                   "Sum Weighted L1I MPKI(Ckp M1)","Sum Weighted L1I MPKI(Ckp gem5)","Total Weighted L1I MPKI(Ckp gem5)",
                   "Sum Weighted L1D CMR(Ckp M1)","Sum Weighted L1D CMR(Ckp gem5)","Total Weighted L1D CMR(Ckp gem5)",
                   "Sum Weighted L1D MPKI(Ckp M1)","Sum Weighted L1D MPKI(Ckp gem5)","Total Weighted L1D MPKI(Ckp gem5)",
                   "Sum Weighted L2 CMR(Ckp M1)","Sum Weighted L2 CMR(Ckp gem5)","Total Weighted L2 CMR(Ckp gem5)",
                   "Sum Weighted L2 MPKI(Ckp M1)","Sum Weighted L2 MPKI(Ckp gem5)","Total Weighted L2 MPKI(Ckp gem5)",
                   "Sum Weighted L3 CMR(Ckp M1)","Sum Weighted L3 CMR(Ckp gem5)","Total Weighted L3 CMR(Ckp gem5)",
                   "Sum Weighted L3 MPKI(Ckp M1)","Sum Weighted L3 MPKI(Ckp gem5)","Total Weighted L3 MPKI(Ckp gem5)",
                   "Credibility(Ckp M1)"
                   ],
                  ["int","500.perlbench_r"],["int","502.gcc_r"],["int","505.mcf_r"],["int","520.omnetpp_r"],["int","523.xalancbmk_r"],["int","525.x264_r"],["int","531.deepsjeng_r"],["int","541.leela_r"],["int","548.exchange2_r"],["int","557.xz_r"],
                  ["fp","503.bwaves_r"],["fp","507.cactuBSSN_r"],["fp","508.namd_r"],["fp","510.parest_r"],["fp","511.povray_r"],["fp","519.lbm_r"],
                  ["fp","521.wrf_r"],["fp","526.blender_r"],["fp","527.cam4_r"],["fp","538.imagick_r"],["fp","544.nab_r"],["fp","549.fotonik3d_r"],["fp","554.roms_r"],["","999.specrand_ir"]]
    for info in initial_data:
        sheet3_cache.append(info)

    # sheet1_CPI
    #  int
    sheet1_CPI.merge_cells("A2:A11")
    # float
    sheet1_CPI.merge_cells("A12:A24")

    # sheet3_cache
    #  int
    sheet3_cache.merge_cells("A2:A11")
    # float
    sheet3_cache.merge_cells("A12:A24")

    # 处理shee2-each ckp
    bm_begin_row=0
    bm_end_row=0
    bm=""
    # print(sheet2_CKPs.max_row)

    # 数据处理
    bm_begin_row=2
    sheet2_CKPs.append(["","","","","","","","",""]) # 最后一行特殊处理，用于识别
    side = Side(border_style="thin", color="000000")
    while bm_end_row<sheet2_CKPs.max_row:
        bm=sheet2_CKPs.cell(bm_begin_row,1).value
        for next_row_idx,next_row_cells in enumerate(sheet2_CKPs[str(bm_begin_row)+":"+str(sheet2_CKPs.max_row)],start=0):
            if next_row_cells[0].value==bm:
                # print("Y",next_row_idx,next_row_cells[0].value)
                bm_end_row=bm_begin_row+next_row_idx
                for col,cell in enumerate(next_row_cells,start=1):
                    if col >2 and col <9 :
                        cell.data_type='float'
                    if col == 9:
                        cell.number_format='0.000%'
                # 写入 CPI相关数据
                # =IFS(E2=0,"",F2=0,"",G2<>0,(H2-G2)/G2)
                if str(sheet2_CKPs.cell(bm_end_row,column_index_from_string('E')).value)=="0" or str(sheet2_CKPs.cell(bm_end_row,column_index_from_string('F')).value)=="0":
                    sheet2_CKPs.cell(bm_end_row,9).value="."
                else:
                    sheet2_CKPs.cell(bm_end_row,9).value="=(H"+str(bm_end_row)+"-G"+str(bm_end_row)+")/G"+str(bm_end_row)
                # 写入cache相关数据
                # if "502" in bm:
                # 获取simpt
                simpt=next_row_cells[2].value

                # print(bm,simpt)

                sheet2_CKPs.cell(1,column_index_from_string("M")).value="L1I CMR(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("N")).value="L1I CMR(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("O")).value="Weighted L1I CMR(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("P")).value="Weighted L1I CMR(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("Q")).value="L1I miss(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("R")).value="L1I miss(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("S")).value="Weighted L1I miss(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("T")).value="Weighted L1I miss(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("U")).value="L1I access(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("V")).value="L1I access(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("W")).value="Weighted L1I access(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("X")).value="Weighted L1I access(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("Y")).value="Weighted L1I MPKI(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("Z")).value="Weighted L1I MPKI(Ckp gem5)"

                sheet2_CKPs.cell(1,column_index_from_string("AA")).value="L1D CMR(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("AB")).value="L1D CMR(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("AC")).value="Weighted L1D CMR(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("AD")).value="Weighted L1D CMR(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("AE")).value="L1D miss(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("AF")).value="L1D miss(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("AG")).value="Weighted L1D miss(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("AH")).value="Weighted L1D miss(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("AI")).value="L1D access(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("AJ")).value="L1D access(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("AK")).value="Weighted L1D access(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("AL")).value="Weighted L1D access(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("AM")).value="Weighted L1D MPKI(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("AN")).value="Weighted L1D MPKI(Ckp gem5)"


                sheet2_CKPs.cell(1,column_index_from_string("AO")).value="L2 CMR(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("AP")).value="L2 CMR(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("AQ")).value="Weighted L2 CMR(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("AR")).value="Weighted L2 CMR(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("AS")).value="L2 miss(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("AT")).value="L2 miss(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("AU")).value="Weighted L2 miss(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("AV")).value="Weighted L2 miss(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("AW")).value="L2 access(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("AX")).value="L2 access(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("AY")).value="Weighted L2 access(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("AZ")).value="Weighted L2 access(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("BA")).value="Weighted L2 MPKI(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("BB")).value="Weighted L2 MPKI(Ckp gem5)"

                sheet2_CKPs.cell(1,column_index_from_string("BC")).value="L3 CMR(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("BD")).value="L3 CMR(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("BE")).value="Weighted L3 CMR(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("BF")).value="Weighted L3 CMR(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("BG")).value="L3 miss(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("BH")).value="L3 miss(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("BI")).value="Weighted L3 miss(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("BJ")).value="Weighted L3 miss(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("BK")).value="L3 access(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("BL")).value="L3 access(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("BM")).value="Weighted L3 access(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("BN")).value="Weighted L3 access(Ckp gem5)"
                sheet2_CKPs.cell(1,column_index_from_string("BO")).value="Weighted L3 MPKI(Ckp M1)"
                sheet2_CKPs.cell(1,column_index_from_string("BP")).value="Weighted L3 MPKI(Ckp gem5)"

                # 抓取并计算M1 ckp中cache miss rate
                # 检查是否存在有效数据
                M1CPI=next_row_cells[4].value
                if M1CPI!="0":
                    L1Icache_miss_rate,L1Icache_miss,L1Icache_access,L1Icache_MPKI,\
                        L1Dcache_miss_rate,L1Dcache_miss,L1Dcache_access,L1Dcache_MPKI,\
                        L2cache_miss_rate,L2cache_miss,L2cache_access,L2cache_MPKI,\
                        L3cache_miss_rate,L3cache_miss,L3cache_access,L3cache_MPKI=grep_cache_data_M1(bm,simpt)


                    # sheet2_CKPs.cell(1,column_index_from_string("M")).value="L1I CMR(Ckp M1)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('M')).value=L1Icache_miss_rate
                    # sheet2_CKPs.cell(1,column_index_from_string("O")).value="Weighted L1I CMR(Ckp M1)"
                    if L1Icache_miss_rate!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('O')).value=f"=D{bm_end_row}*M{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("Q")).value="L1I miss(Ckp M1)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('Q')).value=L1Icache_miss
                    # sheet2_CKPs.cell(1,column_index_from_string("S")).value="Weighted L1I miss(Ckp M1)"
                    if L1Icache_miss!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('S')).value=f"=D{bm_end_row}*Q{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("U")).value="L1I access(Ckp M1)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('U')).value=L1Icache_access
                    # sheet2_CKPs.cell(1,column_index_from_string("W")).value="Weighted L1I access(Ckp M1)"
                    if L1Icache_access!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('W')).value=f"=D{bm_end_row}*U{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("Y")).value="Weighted L1I MPKI(Ckp M1)"
                    if L1Icache_MPKI!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('Y')).value=f"=D{bm_end_row}*{L1Icache_MPKI}"

                    # sheet2_CKPs.cell(1,column_index_from_string("AA")).value="L1D CMR(Ckp M1)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('AA')).value=L1Dcache_miss_rate
                    # sheet2_CKPs.cell(1,column_index_from_string("AC")).value="Weighted L1D CMR(Ckp M1)"
                    if L1Dcache_miss_rate!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('AC')).value=f"=D{bm_end_row}*AA{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("AE")).value="L1D miss(Ckp M1)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('AE')).value=L1Dcache_miss
                    # sheet2_CKPs.cell(1,column_index_from_string("AG")).value="Weighted L1D miss(Ckp M1)"
                    if L1Dcache_miss!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('AG')).value=f"=D{bm_end_row}*AE{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("AI")).value="L1D access(Ckp M1)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('AI')).value=L1Dcache_access
                    # sheet2_CKPs.cell(1,column_index_from_string("AK")).value="Weighted L1D access(Ckp M1)"
                    if L1Dcache_access!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('AK')).value=f"=D{bm_end_row}*AI{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("AM")).value="Weighted L1D MPKI(Ckp M1)"
                    if L1Dcache_MPKI!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('AM')).value=f"=D{bm_end_row}*{L1Dcache_MPKI}"

                    # sheet2_CKPs.cell(1,column_index_from_string("AO")).value="L2 CMR(Ckp M1)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('AO')).value=L2cache_miss_rate
                    # sheet2_CKPs.cell(1,column_index_from_string("AQ")).value="Weighted L2 CMR(Ckp M1)"
                    if L2cache_miss_rate!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('AQ')).value=f"=D{bm_end_row}*AO{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("AS")).value="L2 miss(Ckp M1)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('AS')).value=L2cache_miss
                    # sheet2_CKPs.cell(1,column_index_from_string("AU")).value="Weighted L2 miss(Ckp M1)"
                    if L2cache_miss!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('AU')).value=f"=D{bm_end_row}*AS{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("AW")).value="L2 access(Ckp M1)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('AW')).value=L2cache_access
                    # sheet2_CKPs.cell(1,column_index_from_string("AY")).value="Weighted L2 access(Ckp M1)"
                    if L2cache_access!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('AY')).value=f"=D{bm_end_row}*AW{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("BA")).value="Weighted L2 MPKI(Ckp M1)"
                    if L2cache_MPKI!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('BA')).value=f"=D{bm_end_row}*{L2cache_MPKI}"

                    # sheet2_CKPs.cell(1,column_index_from_string("BC")).value="L3 CMR(Ckp M1)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('BC')).value=L3cache_miss_rate
                    # sheet2_CKPs.cell(1,column_index_from_string("BE")).value="Weighted L3 CMR(Ckp M1)"
                    if L3cache_miss_rate!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('BE')).value=f"=D{bm_end_row}*BC{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("BG")).value="L3 miss(Ckp M1)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('BG')).value=L3cache_miss
                    # sheet2_CKPs.cell(1,column_index_from_string("BI")).value="Weighted L3 miss(Ckp M1)"
                    if L3cache_miss!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('BI')).value=f"=D{bm_end_row}*BG{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("BK")).value="L3 access(Ckp M1)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('BK')).value=L3cache_access
                    # sheet2_CKPs.cell(1,column_index_from_string("BM")).value="Weighted L3 access(Ckp M1)"
                    if L3cache_access!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('BM')).value=f"=D{bm_end_row}*BK{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("BO")).value="Weighted L3 MPKI(Ckp M1)"
                    if L3cache_MPKI!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('BO')).value=f"=D{bm_end_row}*{L3cache_MPKI}"

                # 抓取并计算gem5 ckp中cache miss rate
                # 检查是否存在有效数据
                gem5CPI=next_row_cells[5].value

                if gem5CPI!="0":
                    # 获取simpt对应的num
                    num=""
                    with open(f"./runspec_gem5_power/{bm}/{bm}.merge", 'r', encoding='utf-8') as f:
                        for index,line in enumerate(f.readlines(),start=1):
                            m=line.split()
                            if simpt==m[1]:
                                num=index
                                # print(index,simpt)
                                break
                    L1Icache_miss_rate,L1Icache_miss,L1Icache_access,L1Icache_MPKI, \
                        L1Dcache_miss_rate,L1Dcache_miss,L1Dcache_access,L1Dcache_MPKI, \
                        L2cache_miss_rate,L2cache_miss,L2cache_access,L2cache_MPKI, \
                        L3cache_miss_rate,L3cache_miss,L3cache_access,L3cache_MPKI=grep_cache_data_gem5(bm,simpt,num)


                    # sheet2_CKPs.cell(1,column_index_from_string("N")).value="L1I CMR(Ckp gem5)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('N')).value=L1Icache_miss_rate
                     # sheet2_CKPs.cell(1,column_index_from_string("P")).value="Weighted L1I CMR(Ckp gem5)"
                    if L1Icache_miss_rate!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('P')).value=f"=D{bm_end_row}*N{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("R")).value="L1I miss(Ckp gem5)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('R')).value=L1Icache_miss
                    # sheet2_CKPs.cell(1,column_index_from_string("T")).value="Weighted L1I miss(Ckp gem5)"
                    if L1Icache_miss!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('T')).value=f"=D{bm_end_row}*R{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("V")).value="L1I access(Ckp gem5)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('V')).value=L1Icache_access
                    # sheet2_CKPs.cell(1,column_index_from_string("X")).value="Weighted L1I access(Ckp gem5)"
                    if L1Icache_access!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('X')).value=f"=D{bm_end_row}*V{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("Z")).value="Weighted L1I MPKI(Ckp gem5)"
                    if L1Icache_MPKI!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('Z')).value=f"=D{bm_end_row}*{L1Icache_MPKI}"

                    # sheet2_CKPs.cell(1,column_index_from_string("AB")).value="L1D CMR(Ckp gem5)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('AB')).value=L1Dcache_miss_rate
                    # sheet2_CKPs.cell(1,column_index_from_string("AD")).value="Weighted L1D CMR(Ckp gem5)"
                    if L1Dcache_miss_rate!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('AD')).value=f"=D{bm_end_row}*AB{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("AF")).value="L1D miss(Ckp gem5)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('AF')).value=L1Dcache_miss
                    # sheet2_CKPs.cell(1,column_index_from_string("AH")).value="Weighted L1D miss(Ckp gem5)"
                    if L1Dcache_miss!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('AH')).value=f"=D{bm_end_row}*AF{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("AJ")).value="L1D access(Ckp gem5)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('AJ')).value=L1Dcache_access
                    # sheet2_CKPs.cell(1,column_index_from_string("AL")).value="Weighted L1D access(Ckp gem5)"
                    if L1Dcache_access!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('AL')).value=f"=D{bm_end_row}*AJ{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("AN")).value="Weighted L1D MPKI(Ckp gem5)"
                    if L1Dcache_MPKI!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('AN')).value=f"=D{bm_end_row}*{L1Dcache_MPKI}"

                    # sheet2_CKPs.cell(1,column_index_from_string("AP")).value="L2 CMR(Ckp gem5)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('AP')).value=L2cache_miss_rate
                    # sheet2_CKPs.cell(1,column_index_from_string("AR")).value="Weighted L2 CMR(Ckp gem5)"
                    if L2cache_miss_rate!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('AR')).value=f"=D{bm_end_row}*AP{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("AT")).value="L2 miss(Ckp gem5)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('AT')).value=L2cache_miss
                    # sheet2_CKPs.cell(1,column_index_from_string("AV")).value="Weighted L2 miss(Ckp gem5)"
                    if L2cache_miss!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('AV')).value=f"=D{bm_end_row}*AT{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("AX")).value="L2 access(Ckp gem5)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('AX')).value=L2cache_access
                    # sheet2_CKPs.cell(1,column_index_from_string("AZ")).value="Weighted L2 access(Ckp gem5)"
                    if L2cache_access!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('AZ')).value=f"=D{bm_end_row}*AX{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("BB")).value="Weighted L2 MPKI(Ckp gem5)"
                    if L2cache_MPKI!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('BB')).value=f"=D{bm_end_row}*{L2cache_MPKI}"

                    # sheet2_CKPs.cell(1,column_index_from_string("BD")).value="L3 CMR(Ckp gem5)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('BD')).value=L3cache_miss_rate
                    # sheet2_CKPs.cell(1,column_index_from_string("BF")).value="Weighted L3 CMR(Ckp gem5)"
                    if L3cache_miss_rate!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('BF')).value=f"=D{bm_end_row}*BD{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("BH")).value="L3 miss(Ckp gem5)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('BH')).value=L3cache_miss
                    # sheet2_CKPs.cell(1,column_index_from_string("BJ")).value="Weighted L3 miss(Ckp gem5)"
                    if L3cache_miss!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('BJ')).value=f"=D{bm_end_row}*BH{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("BL")).value="L3 access(Ckp gem5)"
                    sheet2_CKPs.cell(bm_end_row,column_index_from_string('BL')).value=L3cache_access
                    # sheet2_CKPs.cell(1,column_index_from_string("BN")).value="Weighted L3 access(Ckp gem5)"
                    if L3cache_access!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('BN')).value=f"=D{bm_end_row}*BL{bm_end_row}"
                    # sheet2_CKPs.cell(1,column_index_from_string("BP")).value="Weighted L3 MPKI(Ckp gem5)"
                    if L3cache_MPKI!=None:
                        sheet2_CKPs.cell(bm_end_row,column_index_from_string('BP')).value=f"=D{bm_end_row}*{L3cache_MPKI}"

            # 进入下一个测例
            else:
                # print("N",next_row_idx,next_row_cells[0].value,"A"+str(bm_begin_row)+":"+"A"+str(bm_end_row))
                # print(bm_end_row,sheet2_CKPs.max_row)
                # merge
                for cell in next_row_cells:
                    # print(cell)
                    cell.border = Border(top=Side(border_style="thin", color="999999"))
                area="A"+str(bm_begin_row)+":"+"A"+str(bm_end_row)
                # sheet2_CKPs.merge_cells(area)
                for sheet1row_num,sheet1row_cells in enumerate(sheet1_CPI.rows,start=1):
                    if sheet1row_cells[1].value == bm:
                        sheet1row_cells[2].value="=SUMIFS(" \
                                                "CKPs!G"+str(bm_begin_row)+":G"+str(bm_end_row)+","+ \
                                                "CKPs!E"+str(bm_begin_row)+":E"+str(bm_end_row)+","+"\"<>0\""+"," \
                                                "CKPs!F"+str(bm_begin_row)+":F"+str(bm_end_row)+","+"\"<>0\""+\
                                                ")"
                        sheet1row_cells[3].value="=SUMIFS(" \
                                                "CKPs!H"+str(bm_begin_row)+":H"+str(bm_end_row)+","+\
                                                "CKPs!E"+str(bm_begin_row)+":E"+str(bm_end_row)+","+"\"<>0\""+","+\
                                                "CKPs!F"+str(bm_begin_row)+":F"+str(bm_end_row)+","+"\"<>0\""+\
                                                ")"
                        sheet1row_cells[4].value="=SUM(CKPs!H"+str(bm_begin_row)+":H"+str(bm_end_row)+")"
                        sheet1row_cells[5].value=\
                            "=(D"+str(sheet1row_num)+"-C"+str(sheet1row_num)+")"+"/C"+str(sheet1row_num)
                        sheet1row_cells[6].value="=SUMIFS(" \
                                                "'CKPs'!D"+str(bm_begin_row)+":D"+str(bm_end_row)+","+ \
                                                "'CKPs'!E"+str(bm_begin_row)+":E"+str(bm_end_row)+","+"\"<>0\""+")"
                        sheet1row_cells[5].number_format='0.000%'
                        sheet1row_cells[6].number_format='0.000%'

                for row_num,row_cells in enumerate(sheet3_cache.rows,start=1):
                    if sheet3_cache.cell(row_num,column_index_from_string('B')).value == bm:
                        # Sum Weighted L1I CMR(Ckp M1)
                        sheet3_cache.cell(row_num,column_index_from_string('C')).value=f"=SUMIFS(" \
                                                                                       f"CKPs!O{bm_begin_row}:O{bm_end_row}," \
                                                                                       f"CKPs!E{bm_begin_row}:E{bm_end_row},\"<>0\"," \
                                                                                       f"CKPs!F{bm_begin_row}:F{bm_end_row},\"<>0\")"
                        # Sum Weighted L1I CMR(Ckp gem5)
                        sheet3_cache.cell(row_num,column_index_from_string('D')).value=f"=SUMIFS(" \
                                                                                       f"CKPs!P{bm_begin_row}:P{bm_end_row}," \
                                                                                       f"CKPs!E{bm_begin_row}:E{bm_end_row},\"<>0\"," \
                                                                                       f"CKPs!F{bm_begin_row}:F{bm_end_row},\"<>0\")"
                        # Total Weighted L1I CMR(Ckp gem5)
                        sheet3_cache.cell(row_num,column_index_from_string('E')).value=f"=SUM(" \
                                                                                       f"CKPs!P{bm_begin_row}:P{bm_end_row})"
                        # Sum Weighted L1I MPKI(Ckp M1)
                        sheet3_cache.cell(row_num,column_index_from_string('F')).value=f"=SUMIFS(" \
                                                                                       f"CKPs!Y{bm_begin_row}:Y{bm_end_row}," \
                                                                                       f"CKPs!E{bm_begin_row}:E{bm_end_row},\"<>0\"," \
                                                                                       f"CKPs!F{bm_begin_row}:F{bm_end_row},\"<>0\")"
                        # Sum Weighted L1I MPKI(Ckp gem5)
                        sheet3_cache.cell(row_num,column_index_from_string('G')).value=f"=SUMIFS(" \
                                                                                       f"CKPs!Z{bm_begin_row}:Z{bm_end_row}," \
                                                                                       f"CKPs!E{bm_begin_row}:E{bm_end_row},\"<>0\"," \
                                                                                       f"CKPs!F{bm_begin_row}:F{bm_end_row},\"<>0\")"
                        # Total Weighted L1I MPKI(Ckp gem5)
                        sheet3_cache.cell(row_num,column_index_from_string('H')).value=f"=SUM(" \
                                                                                       f"CKPs!Z{bm_begin_row}:Z{bm_end_row})"
                        # Sum Weighted L1D CMR(Ckp M1)
                        sheet3_cache.cell(row_num,column_index_from_string('I')).value=f"=SUMIFS(" \
                                                                                       f"CKPs!AC{bm_begin_row}:AC{bm_end_row}," \
                                                                                       f"CKPs!E{bm_begin_row}:E{bm_end_row},\"<>0\"," \
                                                                                       f"CKPs!F{bm_begin_row}:F{bm_end_row},\"<>0\")"
                        # Sum Weighted L1D CMR(Ckp gem5)
                        sheet3_cache.cell(row_num,column_index_from_string('J')).value=f"=SUMIFS(" \
                                                                                       f"CKPs!AD{bm_begin_row}:AD{bm_end_row}," \
                                                                                       f"CKPs!E{bm_begin_row}:E{bm_end_row},\"<>0\"," \
                                                                                       f"CKPs!F{bm_begin_row}:F{bm_end_row},\"<>0\")"
                        # Total Weighted L1D CMR(Ckp gem5)
                        sheet3_cache.cell(row_num,column_index_from_string('K')).value=f"=SUM(" \
                                                                                       f"CKPs!AD{bm_begin_row}:AD{bm_end_row})"
                        # Sum Weighted L1D MPKI(Ckp M1)
                        sheet3_cache.cell(row_num,column_index_from_string('L')).value=f"=SUMIFS(" \
                                                                                       f"CKPs!AM{bm_begin_row}:AM{bm_end_row}," \
                                                                                       f"CKPs!E{bm_begin_row}:E{bm_end_row},\"<>0\"," \
                                                                                       f"CKPs!F{bm_begin_row}:F{bm_end_row},\"<>0\")"
                        # Sum Weighted L1D MPKI(Ckp gem5)
                        sheet3_cache.cell(row_num,column_index_from_string('M')).value=f"=SUMIFS(" \
                                                                                       f"CKPs!AN{bm_begin_row}:AN{bm_end_row}," \
                                                                                       f"CKPs!E{bm_begin_row}:E{bm_end_row},\"<>0\"," \
                                                                                       f"CKPs!F{bm_begin_row}:F{bm_end_row},\"<>0\")"
                        # Total Weighted L1D MPKI(Ckp gem5)
                        sheet3_cache.cell(row_num,column_index_from_string('N')).value=f"=SUM(" \
                                                                                       f"CKPs!AN{bm_begin_row}:AN{bm_end_row})"
                        # Sum Weighted L2 CMR(Ckp M1)
                        sheet3_cache.cell(row_num,column_index_from_string('O')).value=f"=SUMIFS(" \
                                                                                       f"CKPs!AQ{bm_begin_row}:AQ{bm_end_row}," \
                                                                                       f"CKPs!E{bm_begin_row}:E{bm_end_row},\"<>0\"," \
                                                                                       f"CKPs!F{bm_begin_row}:F{bm_end_row},\"<>0\")"
                        # Sum Weighted L2 CMR(Ckp gem5)
                        sheet3_cache.cell(row_num,column_index_from_string('P')).value=f"=SUMIFS(" \
                                                                                       f"CKPs!AR{bm_begin_row}:AR{bm_end_row}," \
                                                                                       f"CKPs!E{bm_begin_row}:E{bm_end_row},\"<>0\"," \
                                                                                       f"CKPs!F{bm_begin_row}:F{bm_end_row},\"<>0\")"
                        # Total Weighted L2 CMR(Ckp gem5)
                        sheet3_cache.cell(row_num,column_index_from_string('Q')).value=f"=SUM(" \
                                                                                       f"CKPs!AR{bm_begin_row}:AR{bm_end_row})"
                        # Sum Weighted L2 MPKI(Ckp M1)
                        sheet3_cache.cell(row_num,column_index_from_string('R')).value=f"=SUMIFS(" \
                                                                                       f"CKPs!BA{bm_begin_row}:BA{bm_end_row}," \
                                                                                       f"CKPs!E{bm_begin_row}:E{bm_end_row},\"<>0\"," \
                                                                                       f"CKPs!F{bm_begin_row}:F{bm_end_row},\"<>0\")"
                        # Sum Weighted L2 MPKI(Ckp gem5)
                        sheet3_cache.cell(row_num,column_index_from_string('S')).value=f"=SUMIFS(" \
                                                                                       f"CKPs!BB{bm_begin_row}:BB{bm_end_row}," \
                                                                                       f"CKPs!E{bm_begin_row}:E{bm_end_row},\"<>0\"," \
                                                                                       f"CKPs!F{bm_begin_row}:F{bm_end_row},\"<>0\")"
                        # Total Weighted L2 MPKI(Ckp gem5)
                        sheet3_cache.cell(row_num,column_index_from_string('T')).value=f"=SUM(" \
                                                                                       f"CKPs!BB{bm_begin_row}:BB{bm_end_row})"
                        # Sum Weighted L3 CMR(Ckp M1)
                        sheet3_cache.cell(row_num,column_index_from_string('U')).value=f"=SUMIFS(" \
                                                                                       f"CKPs!BE{bm_begin_row}:BE{bm_end_row}," \
                                                                                       f"CKPs!E{bm_begin_row}:E{bm_end_row},\"<>0\"," \
                                                                                       f"CKPs!F{bm_begin_row}:F{bm_end_row},\"<>0\")"
                        # Sum Weighted L3 CMR(Ckp gem5)
                        sheet3_cache.cell(row_num,column_index_from_string('V')).value=f"=SUMIFS(" \
                                                                                       f"CKPs!BF{bm_begin_row}:BF{bm_end_row}," \
                                                                                       f"CKPs!E{bm_begin_row}:E{bm_end_row},\"<>0\"," \
                                                                                       f"CKPs!F{bm_begin_row}:F{bm_end_row},\"<>0\")"
                        # Total Weighted L3 CMR(Ckp gem5)
                        sheet3_cache.cell(row_num,column_index_from_string('W')).value=f"=SUM(" \
                                                                                       f"CKPs!BF{bm_begin_row}:BF{bm_end_row})"
                        # Sum Weighted L3 MPKI(Ckp M1)
                        sheet3_cache.cell(row_num,column_index_from_string('X')).value=f"=SUMIFS(" \
                                                                                       f"CKPs!BO{bm_begin_row}:BO{bm_end_row}," \
                                                                                       f"CKPs!E{bm_begin_row}:E{bm_end_row},\"<>0\"," \
                                                                                       f"CKPs!F{bm_begin_row}:F{bm_end_row},\"<>0\")"
                        # Sum Weighted L3 MPKI(Ckp gem5)
                        sheet3_cache.cell(row_num,column_index_from_string('Y')).value=f"=SUMIFS(" \
                                                                                       f"CKPs!BP{bm_begin_row}:BP{bm_end_row}," \
                                                                                       f"CKPs!E{bm_begin_row}:E{bm_end_row},\"<>0\"," \
                                                                                       f"CKPs!F{bm_begin_row}:F{bm_end_row},\"<>0\")"
                        # Total Weighted L3 MPKI(Ckp gem5)
                        sheet3_cache.cell(row_num,column_index_from_string('Z')).value=f"=SUM(" \
                                                                                       f"CKPs!BP{bm_begin_row}:BP{bm_end_row})"
                        # Credibility(Ckp M1)
                        sheet3_cache.cell(row_num,column_index_from_string('AA')).value=f"=SUMIFS(" \
                                                                                       f"'CKPs'!D{bm_begin_row}:D{bm_end_row}," \
                                                                                       f"'CKPs'!E{bm_begin_row}:E{bm_end_row},\"<>0\")"


                        # sheet1row_cells[5].number_format='0.000%'
                        # sheet1row_cells[6].number_format='0.000%'

                bm_begin_row=bm_begin_row+next_row_idx
                bm_end_row=bm_begin_row
                # print(bm_begin_row)
                break

    # 刷新sheet0
    # gem5 total ckp CPI
    for index,row_cells in enumerate(sheet0["N2:N25"],start=2):
        for cell in row_cells:
            cell.value="=summaryCPI!E"+str(index)
    # gem5 sum ckp CPI
    for index,row_cells in enumerate(sheet0["O2:O25"],start=2):
        for cell in row_cells:
            cell.value="=summaryCPI!D"+str(index)

    for index,row_cells in enumerate(sheet0["J2:J25"],start=2):
        for cell in row_cells:
            cell.value="=IFERROR((1/summaryCPI!D"+str(index)+"-G"+str(index)+")/G"+str(index)+",1)"

    # results_compare.writerow(["Benchmark#","Checkpoint#","Simpts","Weights","M1_CPI","gem5_CPI","M1_WeightedCPI","gem5_WeightedCPI"])

    # 样式处理

    # 过滤非0
    sheet2_CKPs.auto_filter.ref = "A1:J"+str(sheet2_CKPs.max_row)
    filter_data=[]
    for cells in sheet2_CKPs["E1"+":"+"E"+str(sheet2_CKPs.max_row)]:
        if cells[0].value!='0':
            filter_data.append(cells[0].value)
    sheet2_CKPs.auto_filter.add_filter_column(column_index_from_string("E")-1, filter_data)
    # 修改标题
    # sheet0.cell(1,column_index_from_string("I")).value="gem5 ckp与完整跑误差"
    # sheet0.cell(1,column_index_from_string("N")).value="GEM5 ckp CPI"
    # sheet0.cell(1,column_index_from_string("O")).value="M1 ckp CPI"
    sheet2_CKPs.cell(1,column_index_from_string("E")).value="CPI(Ckp M1)"
    sheet2_CKPs.cell(1,column_index_from_string("F")).value="CPI(Ckp gem5)"
    sheet2_CKPs.cell(1,column_index_from_string("G")).value="WeightedCPI(Ckp M1)"
    sheet2_CKPs.cell(1,column_index_from_string("H")).value="WeightedCPI(Ckp gem5)"
    sheet2_CKPs.cell(1,column_index_from_string("J")).value="Valid"
    sheet2_CKPs.cell(1,column_index_from_string("K")).value="Expect Skipped # vgi binary records"
    sheet2_CKPs.cell(1,column_index_from_string("L")).value="Real Skipped # vgi binary records"

    # 条件格式
    rule3 = FormulaRule(formula=['J2=1'], fill=PatternFill(end_color='0099ff'))
    sheet2_CKPs.conditional_formatting.add("J2:"+"J"+str(sheet2_CKPs.max_row-1),rule3)
    rule2=CellIsRule(operator='between',formula=[-0.25,0.25],fill=PatternFill(end_color='33cc33'))#筛选25%以内的绿色显示
    sheet2_CKPs.conditional_formatting.add("I1:"+"I"+str(sheet2_CKPs.max_row-1),rule2)
    rule3 = FormulaRule(formula=['AND(I2>-0.25,I2<0.25)'], fill=PatternFill(end_color='33cc33'))
    sheet2_CKPs.conditional_formatting.add("B2:"+"B"+str(sheet2_CKPs.max_row-1),rule3)
    rule3 = FormulaRule(formula=['AND(F2>-0.25,F2<0.25)'], fill=PatternFill(end_color='33cc33'))
    sheet1_CPI.conditional_formatting.add("B2:"+"B"+str(sheet1_CPI.max_row-1),rule3)
    sheet1_CPI.conditional_formatting.add("F2:"+"F"+str(sheet1_CPI.max_row-1),rule3)

    # 设置边框、对齐、字体
    for index,row_cells in enumerate(sheet0["E2:E25"],start=2):
        for cell in row_cells:
                cell.alignment=Alignment(horizontal='center', vertical='center')
    for index,row_cells in enumerate(sheet0["H2:H25"],start=2):
        for cell in row_cells:
                cell.alignment=Alignment(horizontal='center', vertical='center')
    for index,row_cells in enumerate(sheet0["I2:I25"],start=2):
        for cell in row_cells:
                cell.alignment=Alignment(horizontal='center', vertical='center')
    pyCellFontStyle([sheet1_CPI,sheet2_CKPs],onlyRow=True,set_border=True)
    pyCellFontStyle([sheet1_CPI,sheet2_CKPs],rowAndCol=True,set_align=True,horizontal='center',vertical='center')
    pyCellFontStyle([sheet1_CPI,sheet2_CKPs],rowAndCol=True,set_align=True,rowNumStart=2,colNumStart=2,horizontal='justify')
    pyCellFontStyle([sheet2_CKPs],rowAndCol=True,rowNumStart=2,colNumStart=column_index_from_string("I"),set_align=True,horizontal='right',vertical='justify')
    pyCellFontStyle([sheet1_CPI],rowAndCol=True,rowNumStart=2,colNumStart=column_index_from_string("F"),set_align=True,horizontal='right',vertical='justify')
    pyCellFontStyle([sheet1_CPI],onlyRow=True,rowNumStart=11,set_border=True,border_bottom=True)
    pyCellFontStyle([sheet1_CPI,sheet2_CKPs],onlyRow=True,isbold=True)
    for col in sheet1_CPI.columns:
        col[10].border = Border(bottom= Side(border_style="medium", color="999999"))
        col[sheet1_CPI.max_row-1].border = Border(bottom = Side(border_style="medium", color="999999"))
    # 调整宽度
    auto_set_column_width(sheet1_CPI)
    auto_set_column_width(sheet2_CKPs)
    auto_set_column_width(sheet3_cache)
    sheet2_CKPs.column_dimensions["B"].width = sheet2_CKPs.column_dimensions["B"].width + 2
    sheet2_CKPs.column_dimensions["C"].width = sheet2_CKPs.column_dimensions["C"].width + 2
    sheet2_CKPs.column_dimensions["D"].width = sheet2_CKPs.column_dimensions["D"].width + 2
    sheet2_CKPs.column_dimensions["E"].width = sheet2_CKPs.column_dimensions["E"].width + 2
    sheet2_CKPs.column_dimensions["F"].width = sheet2_CKPs.column_dimensions["F"].width + 2
    sheet2_CKPs.column_dimensions["G"].width = sheet2_CKPs.column_dimensions["G"].width + 2
    sheet2_CKPs.column_dimensions["H"].width = sheet2_CKPs.column_dimensions["H"].width + 2
    sheet2_CKPs.column_dimensions["I"].width = sheet2_CKPs.column_dimensions["I"].width + 2
    # 冻结首行, A2是因为会冻结选定单元格左边的所有列和上面的所有行
    sheet1_CPI.freeze_panes = 'A2'
    sheet2_CKPs.freeze_panes = 'A2'
    workbook.active=sheet0
    workbook.save(gen_file)
    # print("#")
    print("==================================================================================================")
    print("* result Excel:")
    print(os.getcwd()+"/data/gem5/"+begin_time+"-comparison_M1_gem5_SPEC2017_sampling_results" + ".xlsx")
    print("==================================================================================================")
    # print("")
    # print(sheet1_CPI.max_column)
data_pre()
gen_cmp_results()