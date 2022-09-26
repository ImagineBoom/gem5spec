import csv
import re
import os
import subprocess
from pathlib import Path
from collections import defaultdict

import openpyxl
from pandas import read_csv
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Border, Side, Font
# # 根据列的数字返回字母
# print(get_column_letter(2))  # B
# # 根据字母返回列的数字
# print(column_index_from_string('D'))  # 4



exelatency_excel_path=Path("../data/meta/20220908-p8_instr.xlsx")
# 根据实际情况修改gem5的路径
gem5_path=Path("../../gem5-exe-latency")


workbook = load_workbook(filename=exelatency_excel_path)
sheet1=workbook.active


def match_opclass_and_excel():
    pass

def hanlder():
    with open(f'{gem5_path}/src/cpu/p8/O3Execution/P8FuncUnitConfig.py', 'r+', encoding='utf-8') as f:
        filelines=f.readlines()
        isLSU=False
        isLU=False
        f.seek(0,0)
        # 遍历opClass
        for line in filelines:
            group_unit=re.search(
                r'class (?P<unit>\w+)\(FUDesc\):', line
            )
            if group_unit:
                if group_unit["unit"] == "LU":
                    isLU=True
                elif group_unit["unit"] == "LSU":
                    isLSU=True
                else:
                    isLSU=False
                    isLU=False
            group_opClass = re.search(
                r'(?P<prefixSpace>\s+)OpDesc\(opClass="(?P<opClass>\w+)"\s*,\s*opLat=(?P<opLat>\d+)\s*,\s*pipelined=(?P<pipelined>\w+)\),', line
            )

            if group_opClass:
                # 遍历Excel
                for linenum,row in enumerate(sheet1.rows, start=1):
                    MNEMONIC=str(row[0].value)

                    # print(MNEMONIC.capitalize())
                    if group_opClass["opClass"]==MNEMONIC.capitalize():
                        EXE_CYCLES=str(row[2].value)
                        group_excel=re.compile(r'(?P<opLat>\d+)\(.*\)').match(EXE_CYCLES)
                        if group_excel:
                            EXE_CYCLES=group_excel["opLat"]
                            if isLSU or isLU:
                                EXE_CYCLES=f'{EXE_CYCLES} + 2'
                            line=f'{group_opClass["prefixSpace"]}OpDesc(opClass="{group_opClass["opClass"]}", opLat={EXE_CYCLES}, pipelined={group_opClass["pipelined"]}),\n'
                            print(format(linenum,"<10"),format(line))
                            break

            f.write(line)
            # print(group_opClass["opClass"], group_opClass["opLat"], group_opClass["pipelined"])

hanlder()